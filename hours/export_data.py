from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Font, Side
from hours.models import EduPeriod, CourseHours, HoursSettings, \
    SupervisionHours, PracticeHours, OtherHours, CathedraEmployee


def kup_header(ws, teacher):
    year = EduPeriod.objects.filter(active=True).first()
    cathedra = teacher.cathedra.name
    position = teacher.position.name
    teacher_name = str(teacher.FIO).split()
    employee = CathedraEmployee.objects.filter(teacher_id=teacher.id).first()
    ws.cell(row=1, column=1).value = teacher_name[0][0] + teacher_name[1][0] + teacher_name[2][0]
    ws.cell(row=2, column=1).value = 'КАРТОЧКА УЧЕБНЫХ ПОРУЧЕНИЙ на ' + str(year) + ' учебный год'
    ws.cell(row=3, column=1).value = '(плановый набор, первая половина рабочего дня)'
    ws.cell(row=4, column=1).value = 'Кафедра: ' + str(cathedra)
    ws.cell(row=5, column=1).value = 'Преподаватель: ' + str(teacher.FIO)
    ws.cell(row=5, column=13).value = 'Должность: ' + str(position) + ', ставка: ' + str(
        employee.stavka) + ' став., ' + ' уч. степень: '
    ws.cell(row=6, column=1).value = '№'
    ws.cell(row=6, column=2).value = 'Индекс дисциплины'
    ws.cell(row=6, column=3).value = 'Наименование дисциплин'
    ws.cell(row=6, column=4).value = 'курс'
    ws.cell(row=6, column=5).value = 'название группы'
    ws.cell(row=6, column=6).value = 'Студентов'
    ws.cell(row=6, column=7).value = 'Кол-во групп'
    ws.cell(row=6, column=8).value = 'Кол-во подгрупп'
    ws.cell(row=6, column=9).value = 'Лекции'
    ws.cell(row=6, column=10).value = 'Практ., семин. занятия'
    ws.cell(row=7, column=10).value = 'по плану'
    ws.cell(row=7, column=11).value = 'всего'
    ws.cell(row=6, column=12).value = 'Лабор. занятия'
    ws.cell(row=7, column=12).value = 'по плану'
    ws.cell(row=7, column=13).value = 'всего'
    ws.cell(row=6, column=14).value = 'консультации'
    ws.cell(row=7, column=14).value = 'предэкзаменационные'
    ws.cell(row=6, column=15).value = 'прием'
    ws.cell(row=7, column=15).value = 'Зачетов'
    ws.cell(row=7, column=16).value = 'Экзаменов'
    ws.cell(row=6, column=17).value = 'Проверка, РГР, реф., контрольных работ'
    ws.cell(row=6, column=18).value = 'руководство'
    ws.cell(row=7, column=18).value = 'дипл. проектир'
    ws.cell(row=7, column=19).value = 'курс. работами'
    ws.cell(row=7, column=20).value = 'аспирантами'
    ws.cell(row=7, column=21).value = 'магистрантами'
    ws.cell(row=7, column=22).value = 'программой магистратуры'
    ws.cell(row=6, column=23).value = 'руководство практиками'
    ws.cell(row=7, column=23).value = 'учебная'
    ws.cell(row=7, column=24).value = 'производственная'
    ws.cell(row=7, column=25).value = 'преддипломная'
    ws.cell(row=7, column=26).value = 'педагогическая'
    ws.cell(row=6, column=27).value = 'ГЭК, ГАК'
    ws.cell(row=6, column=28).value = 'Проверка остаточных знаний'
    ws.cell(row=6, column=29).value = 'Рецензирование дипл. работ'
    ws.cell(row=6, column=30).value = 'Прием экзаменов по канд. минимуму'
    ws.cell(row=6, column=31).value = 'Рецензир. рефератов для канд. мин.'
    ws.cell(row=6, column=32).value = 'Контроль СРС'
    ws.cell(row=6, column=33).value = 'Контроль БРС'
    ws.cell(row=6, column=34).value = 'Всего часов к расчету штатов'
    ws.cell(row=6, column=35).value = 'Почасовой фонд (для сторонних)'
    ws.cell(row=6, column=36).value = 'ИТОГО по кафедре'
    for k in range(1, 37):
        ws.cell(row=8, column=k).value = k
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=36)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=36)
    ws.merge_cells(start_row=6, start_column=1, end_row=7, end_column=1)
    ws.merge_cells(start_row=6, start_column=2, end_row=7, end_column=2)
    ws.merge_cells(start_row=6, start_column=3, end_row=7, end_column=3)
    ws.merge_cells(start_row=6, start_column=4, end_row=7, end_column=4)
    ws.merge_cells(start_row=6, start_column=5, end_row=7, end_column=5)
    ws.merge_cells(start_row=6, start_column=6, end_row=7, end_column=6)
    ws.merge_cells(start_row=6, start_column=7, end_row=7, end_column=7)
    ws.merge_cells(start_row=6, start_column=8, end_row=7, end_column=8)
    ws.merge_cells(start_row=6, start_column=9, end_row=7, end_column=9)
    ws.merge_cells(start_row=6, start_column=10, end_row=6, end_column=11)
    ws.merge_cells(start_row=6, start_column=12, end_row=6, end_column=13)
    ws.merge_cells(start_row=6, start_column=15, end_row=6, end_column=16)
    ws.merge_cells(start_row=6, start_column=17, end_row=7, end_column=17)
    ws.merge_cells(start_row=6, start_column=18, end_row=6, end_column=22)
    ws.merge_cells(start_row=6, start_column=23, end_row=6, end_column=26)
    ws.merge_cells(start_row=6, start_column=27, end_row=7, end_column=27)
    ws.merge_cells(start_row=6, start_column=28, end_row=7, end_column=28)
    ws.merge_cells(start_row=6, start_column=29, end_row=7, end_column=29)
    ws.merge_cells(start_row=6, start_column=30, end_row=7, end_column=30)
    ws.merge_cells(start_row=6, start_column=31, end_row=7, end_column=31)
    ws.merge_cells(start_row=6, start_column=32, end_row=7, end_column=32)
    ws.merge_cells(start_row=6, start_column=33, end_row=7, end_column=33)
    ws.merge_cells(start_row=6, start_column=34, end_row=7, end_column=34)
    ws.merge_cells(start_row=6, start_column=35, end_row=7, end_column=35)
    ws.merge_cells(start_row=6, start_column=36, end_row=7, end_column=36)
    ws.row_dimensions[6].height = float(33.75)
    ws.row_dimensions[7].height = float(150.00)
    ws.column_dimensions['C'].width = float(50.00)
    ws.column_dimensions['E'].width = float(20.00)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=3, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=6, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=6, column=2).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=3).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=6, column=4).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=5).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=6).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=7).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=8).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=9).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=10).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    ws.cell(row=6, column=12).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    ws.cell(row=6, column=14).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    ws.cell(row=6, column=15).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    ws.cell(row=6, column=18).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    ws.cell(row=6, column=23).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    ws.cell(row=7, column=10).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=7, column=11).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=7, column=12).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=7, column=13).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=7, column=14).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=7, column=15).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=7, column=16).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=17).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=7, column=18).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=7, column=19).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=7, column=20).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=7, column=21).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=7, column=22).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=7, column=23).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=7, column=24).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=7, column=25).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=7, column=26).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=27).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=28).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=29).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=30).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=31).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=32).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=33).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=34).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=35).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    ws.cell(row=6, column=36).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
    for i in range(1, 37):
        ws.cell(row=8, column=i).alignment = Alignment(horizontal='center', vertical='center')
    for i in range(1, 6):
        ws.cell(row=i, column=1).font = Font(bold=True)
    for j in range(2, 6):
        ws.cell(row=6, column=j).font = Font(bold=True)
    for j in range(27, 37):
        ws.cell(row=6, column=j).font = Font(bold=True)
    for j in range(1, 37):
        ws.cell(row=8, column=j).font = Font(bold=True)
    ws.cell(row=5, column=13).font = Font(bold=True)


def kup_body(ws, teacher):
    start = 9
    row_data = 0
    year = EduPeriod.objects.filter(active=True).first()
    courses = CourseHours.objects.filter(teacher_id=teacher.id, edu_period=year)
    settings = HoursSettings.objects.filter(is_active=True).first()
    ws.cell(row=start, column=2).value = 'Первый семестр'
    ws.merge_cells(start_row=start, start_column=2, end_row=start, end_column=36)
    ws.cell(row=start, column=2).font = Font(bold=True)
    for course in courses:
        supervision = SupervisionHours.objects.filter(teacher_id=teacher.id, edu_period=year,
                                                      group=course.group)
        practice = PracticeHours.objects.filter(teacher_id=teacher.id, edu_period=year, group=course.group)
        other = OtherHours.objects.filter(teacher_id=teacher.id, edu_period=year, group=course.group)
        if int(course.discipline_settings.semester.name) % 2 != 0:
            if course.group.edu_type == 1:
                start = start + 1
                row_data += 1
                ws.cell(row=start, column=2).value = 'Очная форма обучения'
                ws.merge_cells(start_row=start, start_column=2, end_row=start, end_column=36)
                ws.cell(row=start, column=2).font = Font(bold=True)
                kup_body_data(ws, start, course, supervision, practice, other)
                start += 1
                row_data += 1
            if course.group.edu_type == 2:
                start += 1
                row_data += 1
                ws.cell(row=start, column=2).value = 'Заочная форма обучения'
                ws.cell(row=start, column=2).font = Font(bold=True)
                ws.merge_cells(start_row=start, start_column=2, end_row=start, end_column=36)
                kup_body_data(ws, start, course, supervision, practice, other)
                start += 1
                row_data += 1
            if course.group.edu_type == 3:
                start += 1
                row_data += 1
                ws.cell(row=start, column=2).value = 'Очно-заочная форма обучения'
                ws.cell(row=start, column=2).font = Font(bold=True)
                ws.merge_cells(start_row=start, start_column=2, end_row=start, end_column=36)
                kup_body_data(ws, start, course, supervision, practice, other)
                start += 1
                row_data += 1
    start += 1
    sum_row1 = start
    ws.cell(row=start, column=3).value = 'Итого за 1 семестр'
    for j in range(2, 37):
        ws.cell(row=start, column=j).font = Font(bold=True)
    for j in range(1, 37):
        ws.cell(row=start, column=j).fill = PatternFill(start_color='DCDCDC',
                                                        end_color='DCDCDC',
                                                        fill_type='solid')
    if row_data > 0:
        ws.cell(row=start, column=9).value = '=SUM(I' + str(start - row_data) + ':I' + str(start - 1) + ')'
        ws.cell(row=start, column=10).value = '=SUM(J' + str(start - row_data) + ':J' + str(start - 1) + ')'
        ws.cell(row=start, column=11).value = '=SUM(K' + str(start - row_data) + ':K' + str(start - 1) + ')'
        ws.cell(row=start, column=12).value = '=SUM(L' + str(start - row_data) + ':L' + str(start - 1) + ')'
        ws.cell(row=start, column=13).value = '=SUM(M' + str(start - row_data) + ':M' + str(start - 1) + ')'
        ws.cell(row=start, column=14).value = '=SUM(N' + str(start - row_data) + ':N' + str(start - 1) + ')'
        ws.cell(row=start, column=15).value = '=SUM(O' + str(start - row_data) + ':O' + str(start - 1) + ')'
        ws.cell(row=start, column=16).value = '=SUM(P' + str(start - row_data) + ':P' + str(start - 1) + ')'
        ws.cell(row=start, column=17).value = '=SUM(Q' + str(start - row_data) + ':Q' + str(start - 1) + ')'
        ws.cell(row=start, column=18).value = '=SUM(R' + str(start - row_data) + ':R' + str(start - 1) + ')'
        ws.cell(row=start, column=19).value = '=SUM(S' + str(start - row_data) + ':S' + str(start - 1) + ')'
        ws.cell(row=start, column=20).value = '=SUM(T' + str(start - row_data) + ':T' + str(start - 1) + ')'
        ws.cell(row=start, column=21).value = '=SUM(U' + str(start - row_data) + ':U' + str(start - 1) + ')'
        ws.cell(row=start, column=22).value = '=SUM(V' + str(start - row_data) + ':V' + str(start - 1) + ')'
        ws.cell(row=start, column=23).value = '=SUM(W' + str(start - row_data) + ':W' + str(start - 1) + ')'
        ws.cell(row=start, column=24).value = '=SUM(X' + str(start - row_data) + ':X' + str(start - 1) + ')'
        ws.cell(row=start, column=25).value = '=SUM(Y' + str(start - row_data) + ':Y' + str(start - 1) + ')'
        ws.cell(row=start, column=27).value = '=SUM(AA' + str(start - row_data) + ':AA' + str(start - 1) + ')'
        ws.cell(row=start, column=28).value = '=SUM(AB' + str(start - row_data) + ':AB' + str(start - 1) + ')'
        ws.cell(row=start, column=29).value = '=SUM(AC' + str(start - row_data) + ':AC' + str(start - 1) + ')'
        ws.cell(row=start, column=30).value = '=SUM(AD' + str(start - row_data) + ':AD' + str(start - 1) + ')'
        ws.cell(row=start, column=31).value = '=SUM(AE' + str(start - row_data) + ':AE' + str(start - 1) + ')'
        ws.cell(row=start, column=32).value = '=SUM(AF' + str(start - row_data) + ':AF' + str(start - 1) + ')'
        ws.cell(row=start, column=33).value = '=SUM(AG' + str(start - row_data) + ':AG' + str(start - 1) + ')'
        ws.cell(row=start, column=34).value = '=SUM(AH' + str(start - row_data) + ':AH' + str(start - 1) + ')'
        ws.cell(row=start, column=35).value = '=SUM(AI' + str(start - row_data) + ':AI' + str(start - 1) + ')'
        ws.cell(row=start, column=36).value = '=SUM(AJ' + str(start - row_data) + ':AJ' + str(start - 1) + ')'
    else:
        for j in range (9, 37):
            ws.cell(row=start, column=j).value = 0
    start += 2
    row_data = 0
    ws.cell(row=start, column=2).value = 'Второй семестр'
    ws.merge_cells(start_row=start, start_column=2, end_row=start, end_column=36)
    ws.cell(row=start, column=2).font = Font(bold=True)
    for course in courses:
        if int(course.discipline_settings.semester.name) % 2 == 0:
            if course.group.edu_type == 1:
                start += 1
                row_data += 1
                ws.cell(row=start, column=2).value = 'Очная форма обучения'
                ws.merge_cells(start_row=start, start_column=2, end_row=start, end_column=36)
                ws.cell(row=start, column=2).font = Font(bold=True)
                kup_body_data(ws, start, course, supervision, practice, other)
                start += 1
                row_data += 1
            if course.group.edu_type == 2:
                start += 1
                row_data += 1
                ws.cell(row=start, column=2).value = 'Заочная форма обучения'
                ws.merge_cells(start_row=start, start_column=2, end_row=start, end_column=36)
                ws.cell(row=start, column=2).font = Font(bold=True)
                kup_body_data(ws, start, course, supervision, practice, other)
                start += 1
                row_data += 1
            if course.group.edu_type == 3:
                start += 1
                row_data += 1
                ws.cell(row=start + 1, column=2).value = 'Очно-заочная форма обучения'
                ws.merge_cells(start_row=start, start_column=2, end_row=start, end_column=36)
                ws.cell(row=start, column=2).font = Font(bold=True)
                kup_body_data(ws, start, course, supervision, practice, other)
                start += 1
                row_data += 1
    start += 1
    sum_row2 = start
    ws.cell(row=start, column=3).value = 'Итого за 2 семестр'
    for j in range(2, 37):
        ws.cell(row=start, column=j).font = Font(bold=True)
    for j in range(1, 37):
        ws.cell(row=start, column=j).fill = PatternFill(start_color='DCDCDC',
                                                        end_color='DCDCDC',
                                                        fill_type='solid')
    if row_data > 0:
        ws.cell(row=start, column=9).value = '=SUM(I' + str(start - row_data) + ':I' + str(start - 1) + ')'
        ws.cell(row=start, column=10).value = '=SUM(J' + str(start - row_data) + ':J' + str(start - 1) + ')'
        ws.cell(row=start, column=11).value = '=SUM(K' + str(start - row_data) + ':K' + str(start - 1) + ')'
        ws.cell(row=start, column=12).value = '=SUM(L' + str(start - row_data) + ':L' + str(start - 1) + ')'
        ws.cell(row=start, column=13).value = '=SUM(M' + str(start - row_data) + ':M' + str(start - 1) + ')'
        ws.cell(row=start, column=14).value = '=SUM(N' + str(start - row_data) + ':N' + str(start - 1) + ')'
        ws.cell(row=start, column=15).value = '=SUM(O' + str(start - row_data) + ':O' + str(start - 1) + ')'
        ws.cell(row=start, column=16).value = '=SUM(P' + str(start - row_data) + ':P' + str(start - 1) + ')'
        ws.cell(row=start, column=17).value = '=SUM(Q' + str(start - row_data) + ':Q' + str(start - 1) + ')'
        ws.cell(row=start, column=18).value = '=SUM(R' + str(start - row_data) + ':R' + str(start - 1) + ')'
        ws.cell(row=start, column=19).value = '=SUM(S' + str(start - row_data) + ':S' + str(start - 1) + ')'
        ws.cell(row=start, column=20).value = '=SUM(T' + str(start - row_data) + ':T' + str(start - 1) + ')'
        ws.cell(row=start, column=21).value = '=SUM(U' + str(start - row_data) + ':U' + str(start - 1) + ')'
        ws.cell(row=start, column=22).value = '=SUM(V' + str(start - row_data) + ':V' + str(start - 1) + ')'
        ws.cell(row=start, column=23).value = '=SUM(W' + str(start - row_data) + ':W' + str(start - 1) + ')'
        ws.cell(row=start, column=24).value = '=SUM(X' + str(start - row_data) + ':X' + str(start - 1) + ')'
        ws.cell(row=start, column=25).value = '=SUM(Y' + str(start - row_data) + ':Y' + str(start - 1) + ')'
        ws.cell(row=start, column=27).value = '=SUM(AA' + str(start - row_data) + ':AA' + str(start - 1) + ')'
        ws.cell(row=start, column=28).value = '=SUM(AB' + str(start - row_data) + ':AB' + str(start - 1) + ')'
        ws.cell(row=start, column=29).value = '=SUM(AC' + str(start - row_data) + ':AC' + str(start - 1) + ')'
        ws.cell(row=start, column=30).value = '=SUM(AD' + str(start - row_data) + ':AD' + str(start - 1) + ')'
        ws.cell(row=start, column=31).value = '=SUM(AE' + str(start - row_data) + ':AE' + str(start - 1) + ')'
        ws.cell(row=start, column=32).value = '=SUM(AF' + str(start - row_data) + ':AF' + str(start - 1) + ')'
        ws.cell(row=start, column=33).value = '=SUM(AG' + str(start - row_data) + ':AG' + str(start - 1) + ')'
        ws.cell(row=start, column=34).value = '=SUM(AH' + str(start - row_data) + ':AH' + str(start - 1) + ')'
        ws.cell(row=start, column=35).value = '=SUM(AI' + str(start - row_data) + ':AI' + str(start - 1) + ')'
        ws.cell(row=start, column=36).value = '=SUM(AJ' + str(start - row_data) + ':AJ' + str(start - 1) + ')'
    else:
        for j in range (9, 37):
            ws.cell(row=start, column=j).value = 0
    start += 1
    ws.cell(row=start, column=3).value = 'Всего часов плановых'
    for j in range(2, 37):
        ws.cell(row=start, column=j).font = Font(bold=True)
    for j in range(1, 37):
        ws.cell(row=start, column=j).fill = PatternFill(start_color='DCDCDC',
                                                        end_color='DCDCDC',
                                                        fill_type='solid')
    ws.cell(row=start, column=9).value = '=I' + str(sum_row1) + ' + ' + 'I' + str(sum_row2)
    ws.cell(row=start, column=10).value = '=J' + str(sum_row1) + ' + ' + 'J' + str(sum_row2)
    ws.cell(row=start, column=11).value = '=K' + str(sum_row1) + ' + ' + 'K' + str(sum_row2)
    ws.cell(row=start, column=12).value = '=L' + str(sum_row1) + ' + ' + 'L' + str(sum_row2)
    ws.cell(row=start, column=13).value = '=M' + str(sum_row1) + ' + ' + 'M' + str(sum_row2)
    ws.cell(row=start, column=14).value = '=N' + str(sum_row1) + ' + ' + 'N' + str(sum_row2)
    ws.cell(row=start, column=15).value = '=O' + str(sum_row1) + ' + ' + 'O' + str(sum_row2)
    ws.cell(row=start, column=16).value = '=P' + str(sum_row1) + ' + ' + 'P' + str(sum_row2)
    ws.cell(row=start, column=17).value = '=Q' + str(sum_row1) + ' + ' + 'Q' + str(sum_row2)
    ws.cell(row=start, column=18).value = '=R' + str(sum_row1) + ' + ' + 'R' + str(sum_row2)
    ws.cell(row=start, column=19).value = '=S' + str(sum_row1) + ' + ' + 'S' + str(sum_row2)
    ws.cell(row=start, column=20).value = '=T' + str(sum_row1) + ' + ' + 'T' + str(sum_row2)
    ws.cell(row=start, column=21).value = '=U' + str(sum_row1) + ' + ' + 'U' + str(sum_row2)
    ws.cell(row=start, column=22).value = '=V' + str(sum_row1) + ' + ' + 'V' + str(sum_row2)
    ws.cell(row=start, column=23).value = '=W' + str(sum_row1) + ' + ' + 'W' + str(sum_row2)
    ws.cell(row=start, column=24).value = '=X' + str(sum_row1) + ' + ' + 'X' + str(sum_row2)
    ws.cell(row=start, column=25).value = '=Y' + str(sum_row1) + ' + ' + 'Y' + str(sum_row2)
    ws.cell(row=start, column=27).value = '=AA' + str(sum_row1) + ' + ' + 'AA' + str(sum_row2)
    ws.cell(row=start, column=28).value = '=AB' + str(sum_row1) + ' + ' + 'AB' + str(sum_row2)
    ws.cell(row=start, column=29).value = '=AC' + str(sum_row1) + ' + ' + 'AC' + str(sum_row2)
    ws.cell(row=start, column=30).value = '=AD' + str(sum_row1) + ' + ' + 'AD' + str(sum_row2)
    ws.cell(row=start, column=31).value = '=AE' + str(sum_row1) + ' + ' + 'AE' + str(sum_row2)
    ws.cell(row=start, column=32).value = '=AF' + str(sum_row1) + ' + ' + 'AF' + str(sum_row2)
    ws.cell(row=start, column=33).value = '=AG' + str(sum_row1) + ' + ' + 'AG' + str(sum_row2)
    ws.cell(row=start, column=34).value = '=AH' + str(sum_row1) + ' + ' + 'AH' + str(sum_row2)
    ws.cell(row=start, column=35).value = '=AI' + str(sum_row1) + ' + ' + 'AI' + str(sum_row2)
    ws.cell(row=start, column=36).value = '=AJ' + str(sum_row1) + ' + ' + 'AJ' + str(sum_row2)
    start += 2
    teacher_name = str(teacher.FIO).split()
    zav_cathedra_name = str(settings.zav_cathedra).split()
    ws.cell(row=start, column=3).value = 'Преподаватель'
    ws.cell(row=start, column=6).value = '/' + teacher_name[1][0] + '.' + teacher_name[2][0] + '. ' + teacher_name[0]
    ws.cell(row=start, column=15).value = 'Зав. кафедрой'
    ws.cell(row=start, column=21).value = '/' + zav_cathedra_name[1][0] + '.' + zav_cathedra_name[2][0] + '. ' + zav_cathedra_name[0]
    ws.cell(row=start, column=27).value = 'Дата:'
    ws.cell(row=start, column=3).alignment = Alignment(horizontal='right')
    ws.cell(row=start, column=4).border = Border(bottom=Side(style='thin'))
    ws.cell(row=start, column=5).border = Border(bottom=Side(style='thin'))
    ws.cell(row=start, column=6).border = Border(bottom=Side(style='thin'))
    ws.cell(row=start, column=7).border = Border(bottom=Side(style='thin'))
    ws.cell(row=start, column=8).border = Border(bottom=Side(style='thin'))
    ws.cell(row=start, column=18).border = Border(bottom=Side(style='thin'))
    ws.cell(row=start, column=19).border = Border(bottom=Side(style='thin'))
    ws.cell(row=start, column=20).border = Border(bottom=Side(style='thin'))
    ws.cell(row=start, column=21).border = Border(bottom=Side(style='thin'))
    ws.cell(row=start, column=22).border = Border(bottom=Side(style='thin'))
    ws.cell(row=start, column=23).border = Border(bottom=Side(style='thin'))
    ws.cell(row=start, column=28).border = Border(bottom=Side(style='thin'))
    ws.cell(row=start, column=29).border = Border(bottom=Side(style='thin'))
    ws.cell(row=start, column=30).border = Border(bottom=Side(style='thin'))
    ws.cell(row=start, column=31).border = Border(bottom=Side(style='thin'))
    for i in range(6, start - 1):
        for j in range(1, 37):
            ws.cell(row=i, column=j).border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                                     left=Side(style='thin'), right=Side(style='thin'))


def kup_body_data(ws, start, course, supervision, practice, other):
    start += 1
    ws.cell(row=start, column=2).value = course.discipline_settings.discipline.code
    ws.cell(row=start, column=3).value = course.discipline_settings.discipline.Name
    ws.cell(row=start, column=4).value = course.group.group.year
    ws.cell(row=start, column=5).value = course.group.group.Name
    ws.cell(row=start, column=6).value = course.group.amount
    ws.cell(row=start, column=7).value = 1
    ws.cell(row=start, column=8).value = course.group.subgroup
    ws.cell(row=start, column=9).value = course.f_lecture
    ws.cell(row=start, column=10).value = course.discipline_settings.Practice
    ws.cell(row=start, column=11).value = course.f_practice
    ws.cell(row=start, column=12).value = course.discipline_settings.Lab
    ws.cell(row=start, column=13).value = course.f_lab
    ws.cell(row=start, column=14).value = course.f_consult_hours
    ws.cell(row=start, column=17).value = course.f_control_hours
    if supervision:
        supervision0 = supervision.filter(supervision_type=1).first()
        supervision1 = supervision.filter(supervision_type=2).first()
        supervision2 = supervision.filter(supervision_type=3).first()
        supervision3 = supervision.filter(supervision_type=4).first()
        supervision4 = supervision.filter(supervision_type=5).first()
        ws.cell(row=start, column=18).value = int(supervision0.hours)
        ws.cell(row=start, column=19).value = int(supervision1.hours)
        ws.cell(row=start, column=20).value = int(supervision2.hours)
        ws.cell(row=start, column=21).value = int(supervision3.hours)
        ws.cell(row=start, column=22).value = int(supervision4.hours)
    if practice:
        practice0 = practice.filter(practice_type=1).first()
        practice1 = practice.filter(practice_type=2).first()
        practice2 = practice.filter(practice_type=3).first()
        practice3 = practice.filter(practice_type=4).first()
        ws.cell(row=start, column=23).value = int(practice0.hours)
        ws.cell(row=start, column=24).value = int(practice1.hours)
        ws.cell(row=start, column=25).value = int(practice2.hours)
        ws.cell(row=start, column=26).value = int(practice3.hours)
    if other:
        other0 = other.filter(other_type=1).first()
        other1 = other.filter(other_type=2).first()
        other2 = other.filter(other_type=3).first()
        other3 = other.filter(other_type=4).first()
        ws.cell(row=start, column=27).value = int(other0.hours)
        ws.cell(row=start, column=29).value = int(other1.hours)
        ws.cell(row=start, column=30).value = int(other2.hours)
        ws.cell(row=start, column=31).value = int(other3.hours)
    ws.cell(row=start, column=32).value = course.f_control_SRS
    ws.cell(row=start, column=33).value = course.f_control_BRS
    ws.cell(row=start,
            column=34).value = course.f_lecture + course.f_practice + course.f_lab + course.f_consult_hours + course.f_control_hours + course.f_control_SRS + course.f_control_BRS
    ws.cell(row=start, column=35).value = 0
    ws.cell(row=start,
            column=36).value = int(ws.cell(row=start, column=34).value) + int(ws.cell(row=start, column=35).value)
    start += 1


def kup_export(teacher):
    wb = Workbook()
    ws = wb.active
    teacher_name = str(teacher.FIO).split()
    ws.title = teacher_name[0][0] + teacher_name[1][0] + teacher_name[2][0]
    kup_header(ws, teacher)
    kup_body(ws, teacher)
    return wb


def schedule_request():
    pass


def attachment():
    pass
