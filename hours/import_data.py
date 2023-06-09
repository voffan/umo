import openpyxl
from django.db import transaction
from hours.models import GroupInfo, EduPeriod, Kafedra, DisciplineSetting, CourseHours, HoursSettings, \
    SupervisionHours, PracticeHours, OtherHours, TeacherGekStatus, StudentsGroup
from umo.models import Discipline, EduProgram, Group, Year, GroupList, Student, EduOrg, Specialization, Profile, \
    Control, DisciplineDetails, Semester
import synch.models as sync_models
import datetime
import math


@transaction.atomic
def import_students(file):
    wb_obj = openpyxl.load_workbook(file, data_only=True)
    sheet_obj_oo = wb_obj.active
    m_row = sheet_obj_oo.max_row
    cols, begin_col = parse_cont_title(sheet_obj_oo)
    for i in range(1, m_row + 1):
        if sheet_obj_oo.cell(row=i, column=2).value == "ИМИ":
            name_group = str(sheet_obj_oo.cell(row=i, column=4).value)
            total = int(str(sheet_obj_oo.cell(row=i, column=28).value))
            edu_type = "ОФО"
            print(name_group, total)
            if total >= 20:
                subgroup_number = 2
            else:
                subgroup_number = 1
            group = Group.objects.filter(Name=name_group).first()
            if group is None:
                continue
            add_contingent(group, name_group, subgroup_number, total, edu_type)

            rf, rsa, d = group_student(sheet_obj_oo, cols, i)
            students_group = StudentsGroup.objects.filter(group__group__Name=name_group).first()
            if students_group is None:
                add_group_student(group, name_group, rf, rsa, d)

    sheet_obj_zo = wb_obj['ЗО без академ']
    m_row = sheet_obj_zo.max_row
    for i in range(1, m_row + 1):
        if sheet_obj_zo.cell(row=i, column=2).value == "ИМИ":
            name_group = str(sheet_obj_zo.cell(row=i, column=4).value)
            total = int(str(sheet_obj_zo.cell(row=i, column=27).value))
            edu_type = "ЗФО"
            print(name_group, total)
            if total >= 20:
                subgroup_number = 2
            else:
                subgroup_number = 1
            group = Group.objects.filter(Name=name_group).first()
            if group is None:
                continue
            add_contingent(group, name_group, subgroup_number, total, edu_type)

            rf, rsa, d = group_student(sheet_obj_zo, cols, i)
            students_group = StudentsGroup.objects.filter(group__group__Name=name_group).first()
            if students_group is None:
                add_group_student(group, name_group, rf, rsa, d)

    sheet_obj_ozo = wb_obj['ОЗО без академ']
    m_row = sheet_obj_ozo.max_row
    for i in range(1, m_row + 1):
        if sheet_obj_ozo.cell(row=i, column=2).value == "ИМИ":
            name_group = str(sheet_obj_ozo.cell(row=i, column=4).value)
            total = int(str(sheet_obj_ozo.cell(row=i, column=15).value))
            edu_type = "ОЗО"
            print(name_group, total)
            if total >= 20:
                subgroup_number = 2
            else:
                subgroup_number = 1
            group = Group.objects.filter(Name=name_group).first()
            if group is None:
                continue
            add_contingent(group, name_group, subgroup_number, total, edu_type)

            rf, rsa, d = group_student(sheet_obj_ozo, cols, i)
            students_group = StudentsGroup.objects.filter(group__group__Name=name_group).first()
            if students_group is None:
                add_group_student(group, name_group, rf, rsa, d)


def add_contingent(group, name_group, subgroup_number, total, edu_type):
    contingent = GroupInfo.objects.filter(group__Name=name_group).first()
    if contingent is None:
        contingent = GroupInfo()
        contingent.group = group
        contingent.group_type = 1
    contingent.subgroup = subgroup_number
    contingent.amount = total
    if edu_type == "ОФО":
        contingent.edu_type = contingent.OFO
    elif edu_type == "ЗФО":
        contingent.edu_type = contingent.ZFO
    elif edu_type == "ОЗО":
        contingent.edu_type = contingent.OZO
    contingent.save()


def get_begin_row_col(sheet_obj):
    m_row = sheet_obj.max_row
    m_col = sheet_obj.max_column
    for i in range(1, m_row):
        for j in range(m_col, 0, -1):
            if sheet_obj.cell(row=i, column=j).value == 'Общий итог':
                begin_row = i + 1
                begin_col = j
                return begin_row, begin_col
    return 0, 0


def parse_cont_title(sheet_obj):
    cols = []
    begin_row, begin_col = get_begin_row_col(sheet_obj)
    j = begin_col
    while j > 0:
        if sheet_obj.cell(row=begin_row, column=j).value == 'Д':
            r = [j]
            r += [j-1, j-2] if sheet_obj.cell(row=begin_row, column=j-1).value == 'Б_РС' else [-1, j-1]
            cols += [r]
        j -= 1
    return cols, begin_col


def group_student(sheet_obj, cols, i):
    rf = 0
    rsa = 0
    d = 0
    for k in cols:
        if sheet_obj.cell(row=i, column=k[2]).value:
            rf = sheet_obj.cell(row=i, column=k[2]).value
            if k[1] > 0 and sheet_obj.cell(row=i, column=k[1]).value:
                rsa = sheet_obj.cell(row=i, column=k[1]).value
            if sheet_obj.cell(row=i, column=k[0]).value:
                d = sheet_obj.cell(row=i, column=k[0]).value
    return rf, rsa, d


def add_group_student(group, name_group, rf, rsa, d):
    student_group = StudentsGroup.objects.filter(group__group__Name=name_group).first()
    if student_group is None:
        group_student_form(group, rf, rsa, d, 0)
        group_student_form(group, rf, rsa, d, 1)
        group_student_form(group, rf, rsa, d, 2)


def group_student_form(group, rf, rsa, d, budget_type):
    student_group = StudentsGroup()
    student_group.group = GroupInfo.objects.filter(group=group).first()
    student_group.course = 1
    student_group.budget_type = budget_type
    if budget_type == 0:
        student_group.amount = rf
    if budget_type == 1:
        student_group.amount = rsa
    if budget_type == 2:
        student_group.amount = d
    if student_group.amount != 0:
        student_group.save()


@transaction.atomic
def import_course(file):
    wb_obj = openpyxl.load_workbook(file, data_only=True)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    for i in range(1, m_row + 1):
        if sheet_obj.cell(row=i, column=3).value == "ИМИ" and sheet_obj.cell(row=i, column=23).value == 1:
            semester = int(str(sheet_obj.cell(row=i, column=7).value))
            name_cathedra = int(str(sheet_obj.cell(row=i, column=22).value))
            name_discipline = str(sheet_obj.cell(row=i, column=4).value)
            value_lecture = int(str(sheet_obj.cell(row=i, column=8).value))
            value_practice = int(str(sheet_obj.cell(row=i, column=10).value))
            value_lab = int(str(sheet_obj.cell(row=i, column=9).value))
            value_control = int(str(sheet_obj.cell(row=i, column=21).value))
            value_SRS = int(str(sheet_obj.cell(row=i, column=11).value))
            ctrl_exam = int(str(sheet_obj.cell(row=i, column=12).value))
            ctrl_zacho = int(str(sheet_obj.cell(row=i, column=13).value))
            ctrl_zach = int(str(sheet_obj.cell(row=i, column=14).value))
            ctrl_kp = int(str(sheet_obj.cell(row=i, column=15).value))
            ctrl_kr = int(str(sheet_obj.cell(row=i, column=16).value))
            ctrl_control = int(str(sheet_obj.cell(row=i, column=17).value))
            ctrl_ref = int(str(sheet_obj.cell(row=i, column=18).value))
            ctrl_rgr = int(str(sheet_obj.cell(row=i, column=19).value))
            code = str(sheet_obj.cell(row=i, column=2).value).split("_")[0]
            if "G" in code:
                code = code[1:]
            if "П" in code:
                code = code[:len(code) - 1]
            code = code[0] + code[1] + '.' + code[2] + code[3] + '.' + code[4] + code[5]
            year = str(sheet_obj.cell(row=i, column=2).value)
            if "G" in year:
                year = year[1:]
            seps = ['_', '-']
            for sep in seps:
                year = year.replace(sep, ' ')
            year = int(str('20' + year.split()[1]))
            # Группы
            groups = Group.objects.filter(program__specialization__code=code, begin_year__year=year)
            if len(groups) < 1:
                groups = synch_group(str(sheet_obj.cell(row=i, column=3).value), code,
                                     year)
            for group in groups:
                # Дисцилпины
                dd = DisciplineSetting.objects.filter(discipline__Name=name_discipline,
                                                      discipline__program__id=group.program.id,
                                                      semester__name=semester).first()
                if dd is None:
                    disc_name = str(sheet_obj.cell(row=i, column=4).value)
                    d = Discipline.objects.filter(Name__icontains=disc_name, program_id=group.program.id).first()
                    if d is None:
                        d = Discipline()
                        d.Name = disc_name
                        d.code = code
                        d.program = group.program
                        d.save()
                        dd = DisciplineSetting()
                        dd.Credit = 0
                        dd.Lecture = value_lecture
                        dd.Practice = value_practice
                        dd.Lab = value_lab
                        dd.KSR = value_control
                        dd.SRS = value_SRS
                        dd.semester = Semester.objects.get(name=semester)
                        dd.discipline = d
                    else:
                        disc_detail = DisciplineDetails.objects.filter(discipline__Name=name_discipline,
                                                                       discipline__program__id=group.program.id,
                                                                       semester__name=semester).first()
                        dd = DisciplineSetting(base_ptr=disc_detail)
                    dd.is_lecture_seperate = False
                    dd.practice_weekly = 0
                    dd.is_hourly = False
                    dd.is_KR_KP_VKR = False
                    dd.is_new = False
                    dd.need_new_RPD = False
                    dd.need_upd_RPD = False
                    dd.save()
                    ctrl = Control()
                    ctrl.discipline_detail = dd
                    if ctrl_exam == 1:
                        ctrl.control_type = 1
                    elif ctrl_zacho == 1:
                        ctrl.control_type = 3
                    elif ctrl_zach == 1:
                        ctrl.control_type = 2
                    elif ctrl_kp == 1:
                        ctrl.control_type = 5
                    elif ctrl_kr == 1:
                        ctrl.control_type = 4
                    elif ctrl_control == 1:
                        ctrl.control_type = 6
                    elif ctrl_ref == 1:
                        ctrl.control_type = 10
                    elif ctrl_rgr == 1:
                        ctrl.control_type = 11
                    else:
                        ctrl.control_type = 0
                    ctrl.control_hours = 0
                    ctrl.save()
                # Курсы
                add_course(group, group.cathedra, dd, value_lecture, value_practice, value_lab,
                           value_control, value_SRS)


def synch_group(inst_name, spec_code, begin_year):
    institutes = {"ИМИ": {'id': 1118, 'name': "Институт математики и информатики"},
                  "ФТИ": {'id': 1119, 'name': "Физико-технический институт"},
                  "ПИ": {'id': 1115, 'name': "Педагогический институт"},
                  "ГИ": {'id': 1121, 'name': "Горный институт"},
                  "ИТИ": {'id': 1129, 'name': "Инженерно-технический институт"},
                  "ИЕН": {'id': 1128, 'name': "Институт естественных наук"},
                  "ИЗФИР": {'id': 1122, 'name': "Институт зарубежной филологии и регионоведения"},
                  "ИЯКН": {'id': 1127, 'name': "Институт языков и культуры народов Северо-Востока РФ"},
                  "МИ": {'id': 1116, 'name': "Медицинский институт"},
                  "ФЭИ": {'id': 1117, 'name': "Финансово-экономический институт"},
                  "АДФ": {'id': 1130, 'name': "Автодорожный факультет"},
                  "ГРФ": {'id': 1137, 'name': "Геологоразведочный факультет"},
                  "ИФ": {'id': 1123, 'name': "Исторический факультет"},
                  "ФЛФ": {'id': 1125, 'name': "Филологический факультет"},
                  "ЮФ": {'id': 1126, 'name': "Юридический факультет"},
                  "ИФКиС": {'id': 1120, 'name': "Институт физической культуры и спорта"},
                  "ИП": {'id': 1124, 'name': "Институт психологии"},
                  }
    begin_year = datetime.date(begin_year, 9, 1)
    year = Year.objects.filter(year=begin_year.year).first()
    synch_groups = sync_models.PlnGroupStud.objects.filter(id_pln__id_dop__id_institute=institutes[inst_name]['id'],
                                                           id_pln__id_dop__id_spec__code=spec_code,
                                                           id_pln__datebegin__gte=begin_year.strftime('%Y-%m-%d'))
    n = synch_groups.count()
    i = 1
    groups = []
    for sg in synch_groups:
        print(i, 'of', n)
        eduprogyear = sg.id_pln
        g = Group.objects.filter(id=sg.id_group).first()
        if g is None:
            institute = EduOrg.objects.filter(name=sg.id_pln.id_dop.id_institute).first()  ###
            if institute is None:
                institute = add_institute(institutes[inst_name])
            cathedra = Kafedra.objects.filter(number=sg.id_pln.id_dop.id_kafedra).first()
            if cathedra is None:
                cathedra = add_cathedra(institute, sg.id_pln.id_dop.id_cathedra)  # , name_cathedra)
            edu_program = EduProgram.objects.filter(specialization__code=eduprogyear.id_dop.id_spec.code,
                                                    year__year__lte=eduprogyear.year).order_by(
                '-year__year').first()
            if edu_program is None:
                edu_program = add_edu_program(cathedra, spec_code, year, sg)
            g = Group()
            g.id = sg.id_group
            g.program = edu_program
            g.cathedra = cathedra
        g.begin_year = Year.objects.get_or_create(year=eduprogyear.year)[0]
        g.Name = sg.name
        g.save()
        synch_people = sync_models.PeoplePln.objects.filter(id_group=sg.id_group)
        for sp in synch_people:
            gl = GroupList.objects.filter(id=sp.id_peoplepln).first()
            if sp.id_status != 2 and sp.id_status != 6 and sp.id_status != 9:
                if gl is not None:
                    gl.active = False
                    gl.save()
                continue
            if gl is not None:
                st = gl.student
            else:
                gl = GroupList()
                gl.id = sp.id_peoplepln
                st = Student()
                st.id = sp.id_people.id_people
            st.FIO = sp.id_people.fio
            st.save()
            gl.student = st
            gl.group = g
            gl.active = True
            gl.save()
        i += 1
        groups.append(g)
    return groups


def add_institute(inst_name):
    university = EduOrg.objects.filter(
        uni__name__icontains="Северо-Восточный федеральный университет имени М.К. Аммосова").first()
    e = EduOrg()
    e.name = inst_name['name']
    e.uni = university
    e.save()
    return e


def add_cathedra(institute, id_cathedra):
    cathedras = {"1": "Автомобильные дороги и аэродромы", "2": "Акушерство и гинекология", "3": "Алгебра и геометрия",
                 "4": "Химическое отделение",
                 "5": "Нормальной и патологической анатомии, оперативной хирургии с топографической анатомией и "
                      "судебной медицины",
                 "6": "Английская филология", "7": "(реорг., 37) Археология и история Северо-Востока России",
                 "8": "(реорганизована см.157 ) Биологической химии (реорганизована)",
                 "9": "(реорг., 31) Ботаника и мерзлотное лесоведение", "10": "(упразднена) Военного дела",
                 "11": "Возрастная и педагогическая психология", "12": "Восточные языки и страноведение",
                 "13": "(реорг., 37) Всемирная история и этнология",
                 "14": "(реорг., 4) Высокомолекулярные соединения и органическая химия", "15": "Высшая математика",
                 "16": "Эколого-географическое отделение",
                 "17": "Геофизические методы, поиск и разведка месторождений полезных ископаемых",
                 "18": "Гистология и микробиология", "19": "(реорг., 79) Горные машины",
                 "20": "Архитектура и городское строительство", "21": "Физическое воспитание",
                 "22": "Госпитальная терапия, профессиональные болезни и клиническая фармакология",
                 "23": "Госпитальная хиpургия и лучевая диагностика", "24": "Конституционное и муниципальное право",
                 "25": "Гражданское право и процесс", "26": "(реорганизована с 77) Детской хирургии (реорганизована)",
                 "27": "Офис образовательных программ", "28": "Дифференциальные уравнения", "29": "Журналистика",
                 "30": "(реорг.) Зимних видов спорта (реорганизована)", "31": "Биологическое отделение",
                 "32": "(реорганизовано с 122) Инвестиционной деятельности, банковского и страхового дела ("
                       "реорганизована)",
                 "33": "(реорг., 144) Инженерная графика", "34": "Иностpанные языки по гуманитаpным специальностям",
                 "35": "(реорг с 34) Иностранных языков по педагогическим специальностям (реорганизована)",
                 "36": "Иностранные языки по техническим и естественным специальностям",
                 "37": "Всемирная, отечественная история, этнология, археология", "38": "Информационные технологии",
                 "39": "Информатика и вычислительная техника", "40": "Культуpология",
                 "41": "(упразд УС№08 06.05.09, реорг с 40) Культуры и искусства (реорганизована)",
                 "42": "Циклические виды спорта", "43": "Математический анализ",
                 "44": "Математическая экономика и прикладная информатика", "45": "Машиноведение",
                 "46": "Адаптивная физическая культура", "47": "(реорг., 122) Международные экономические отношения",
                 "48": "Экономика и управление развитием территорий", "49": "(реорг., 80) Мерзлотоведение",
                 "50": "(реорг.) Методика пpеподавания иностpанных языков (реорганизована)",
                 "51": "Педагогическое отделение", "52": "История, обществознание и политология",
                 "53": "Методика преподавания математики", "54": "Технология",
                 "55": "Методика преподавания русского языка и литературы", "56": "Методика преподавания физики",
                 "57": "Методика преподавания якутского языка, литературы и национальной культуры",
                 "58": "(закрыта УС№7 от 01.04.09, реорг с 89) Минералогии и петрографии (реорганизована)",
                 "59": "Мас-рестлинг и национальные виды спорта", "60": "Немецкая филология",
                 "61": "Общее языкознание и риторика", "62": "(реорг., 31) Общая биология",
                 "63": "(упразднена ) Общей врачебной практики (реорганизована)",
                 "64": "(реорганизована с 4) Общей и неорганической химии (реорганизована)",
                 "65": "(реорг., 96) Общая психология", "66": "Общая и экспериментальная физика",
                 "67": "Общая хирургия", "68": "Общественное здоровье и здравоохранение, общая гигиена и биоэтика",
                 "69": "(реорг с 68) Основ здорового образа жизни и экологии человека (реорганизована)",
                 "70": "(реорг.) Основы ядерной физики (реорганизована)", "71": "(реорг., 79) Открытые горные работы",
                 "72": "(реорг., 147) Защита в чрезвычайных ситуациях", "73": "Педагогика",
                 "74": "Дошкольное образование",
                 "75": "(реорг.за 115) Педагогики и психологии высшей школы (реорганизована)",
                 "76": "Начальное образование", "77": "Педиатрия и детская хирургия", "78": "Перевод",
                 "79": "Горное дело", "80": "Прикладная геология", "81": "(реорг., 52) Политология",
                 "82": "Прикладная математика", "83": "Производство строительных материалов, изделий и конструкций",
                 "84": "Пропедевтическая и факультетская терапия с эндокринологией и ЛФК",
                 "85": "Пропедевтика детских болезней", "86": "Неврология и психиатрия",
                 "87": "(реорг., 88) Радиотехника и информационные технологии",
                 "88": "Радиофизика и электронные системы", "89": "(реорг., 80) Региональная геология и геоинформатика",
                 "90": "Реклама и связи с общественностью", "91": "Русский язык",
                 "92": "Русская и зарубежная литература", "93": "Русская литература ХХ века и теория литературы",
                 "94": "Северная филология", "95": "Прикладная механика", "96": "Психология и социальные науки",
                 "97": "Социальная педагогика", "98": "Социология и управление персоналом",
                 "99": "(реорг. с 100) Спортивной борьбы (реорганизована)",
                 "100": "Теория и методика спортивных единоборств",
                 "101": "Спортивно-оздоровительный туризм и массовые виды спорта",
                 "102": "Стилистика якутского языка и русско-якутского перевода",
                 "103": "Промышленное и гражданское строительство",
                 "104": "(реорг., 103) Строительные конструкции и проектирование",
                 "105": "Теория и методика физической культуры и безопасности жизнедеятельности",
                 "106": "Теоретическая физика", "107": "Теория и методика обучения информатике",
                 "108": "Теплогазоснабжение и вентиляция", "109": "Теплофизика и теплоэнергетика",
                 "110": "Терапевтическая, хирургическая ортопедическая стоматология и стоматология детского возраста",
                 "111": "Технология обработки драгоценных камней и металлов",
                 "112": "Технология деревообработки и деревянных конструкций",
                 "113": "(реорг., 150) Технология и техника разведки месторождений полезных ископаемых",
                 "114": "Уголовное право и процесс",
                 "115": "(реорг.,  97) Профессиональная педагогика, психология и управление образованием",
                 "116": "Фpанцузская филология", "117": "(реорг с 84) Факультетской терапии (реорганизована)",
                 "118": "Факультетская хирургия, урология, онкология и отоларингология",
                 "119": "Физика материалов и технология сварки", "120": "Нормальная и патологическая физиология",
                 "121": "Философии", "122": "Экономика и финансы", "123": "Фольклор и культура",
                 "124": "(реорг с 134) Фтизиатрии (реорганизована)",
                 "125": "(реорг с 110) Хирургической стоматологии (реорганизована)", "126": "(реорг., 16) Экология",
                 "127": "Менеджмент", "128": "Экономика труда и социальные отношения", "129": "Экономическая теория",
                 "130": "(реорганизована с 66) Физики (реорганизована)", "131": "Электроснабжение",
                 "132": "Якутский язык", "133": "Якутская литература",
                 "134": "Инфекционные болезни, фтизиатрия и дерматовенерология",
                 "135": "Травматология, ортопедия и медицина катастроф", "136": "Резерв 1", "137": "Резерв 2",
                 "138": "Сестринское дело", "139": "(реорг., 48) Бухгалтерский учет, анализ и аудит",
                 "140": "Теория, история государства и права",
                 "141": "Эксплуатация автомобильного транспорта и автосервис", "142": "Североведения",
                 "143": "Фармакология и фармация", "144": "Экспертиза, управление и кадастр недвижимости",
                 "145": "(реорг., 122) Маркетинг и экономика",
                 "146": "(реорг. с 20) Архитектуры и градостроительства (реорганизована)",
                 "147": "Техносферная безопасность", "148": "Многоканальные телекоммуникационные системы",
                 "149": "(реорг., 119) Сварка, диагностика и мониторинг конструкций", "150": "Недропользование",
                 "151": "Международные исследования", "152": "Специальное (дефектологическое) образование",
                 "153": "Внутренние болезни и общеврачебная практика (семейная медицина) (ФПОВ)",
                 "154": "Хирургические болезни и стоматология (ФПОВ)", "155": "Акушерство и гинекология (ФПОВ)",
                 "156": "Анестезиология, реаниматология и интенсивная терапия с курсом медицинской помощи (ФПОВ)",
                 "157": "(реорг., 4) Биологической химии и биотехнологии",
                 "158": "Социально-культурный сервис и туризм", "159": "Русский как иностранный",
                 "160": "Вычислительные технологии", "161": "Эксплуатации и обслуживания информационных систем (КТ)",
                 "162": "Технических дисциплин (КТ)", "163": "Специальных дисциплин (КТ)",
                 "164": "Резерв (реорганизована)", "165": "Предпринимательское право и клиническое обучение",
                 "166": "Арктическое право и право стран Азиатско-Тихоокеанского региона", "167": "Андрагогики (ИНПО)",
                 "168": "Социальной информатики (ИНПО)", "169": "Учебно-методический центр \"Автошкола СВФУ\""}
    k = Kafedra()
    k.number = id_cathedra
    k.name = cathedras[id_cathedra]
    k.institution = institute
    k.save()
    return k


def add_edu_program(cathedra, spec_code, year, sg):
    e = EduProgram()
    e.cathedra = cathedra
    s = Specialization.objects.filter(code=spec_code, year__year=year.year).first()
    if s is None:
        s = Specialization()
        s.name = sg.id_pln.id_dop.id_spec.id_specialn.name
        s.brief_name = sg.id_pln.id_dop.id_spec.id_specialn.brief_name
        s.code = spec_code
        s.qual = get_qualification(sg.id_pln.id_qualify.id_qualifylevel)
        s.level = get_edu_level(sg.id_pln.id_qualify.id_qualifylevel)
        s.save()
        p = Profile()
        p.spec = s
        p.name = sg.id_pln.id_dop.id_spec.name
        p.save()
    e.specialization = s
    e.year = year
    e.save()
    return e


def get_qualification(qual):
    data = {3: 2, 4: 3, 5: 1, 22: 1, 23: 4, 24: 5, 16: 8}
    return data[qual] if qual in data.keys() else 0


def get_edu_level(level):
    data = {8: 1, 12: 1, 14: 1, 15: 1, 19: 1, 20: 1, 9: 0}
    return data[level] if level in data.keys() else 2


def add_course(group, cathedra, dd, value_lecture, value_practice, value_lab,
               value_control, value_SRS):
    settings = HoursSettings.objects.filter(is_active=True).first()
    courseh = CourseHours()
    courseh.edu_period = EduPeriod.objects.get(active=True)
    courseh.teacher = None
    g = GroupInfo.objects.filter(group__id=group.id).first()
    if g is None:
        g = GroupInfo()
        g.group = group
        g.group_type = 1
        g.subgroup = 1
        g.amount = 1
        g.edu_type = 1
        g.save()
    courseh.group = g
    courseh.cathedra = cathedra
    courseh.discipline_settings = dd
    courseh.f_lecture = value_lecture
    courseh.f_practice = value_practice
    courseh.f_lab = value_lab
    courseh.f_consult_hours = settings.consult if dd.control_set.get().control_type == Control.EXAM else 0
    courseh.f_exam_hours = settings.exam * g.amount if dd.control_set.get().control_type == Control.EXAM else 0
    courseh.f_control_hours = value_control
    courseh.f_check_hours = 0
    courseh.f_control_SRS = value_SRS
    courseh.f_control_BRS = math.ceil(g.amount * settings.brs)  # округление в большую сторону
    courseh.save()


def add_supervision_hours(teacher, group, cathedra, supervision_type, superv):
    settings = HoursSettings.objects.filter(is_active=True).first()
    if superv is None:
        superv = SupervisionHours()
    superv.teacher = teacher
    superv.group = group
    superv.students = group.amount
    superv.supervision_type = supervision_type
    course = CourseHours.objects.filter(teacher_id=teacher.id, group_id=group.id).first()
    control = Control.objects.filter(discipline_detail__discipline__id=course.discipline_settings.discipline.id).first()
    if superv.supervision_type == 1:
        if "Руководство ВКР" in course.discipline_settings.discipline.Name:
            if group.group.program.specialization.qual == 4:
                superv.hours = settings.vkr * superv.students
            else:
                superv.hours = 0
        else:
            superv.hours = 0
    elif superv.supervision_type == 2:
        if control:
            if control.control_type == 4 or 5:
                if group.group.program.specialization.qual == 4:
                    if course.discipline_settings.semester.name == "1" or "2":
                        superv.hours = settings.kp_kr_1 * superv.students
                    else:
                        superv.hours = settings.kp_kr_2 * superv.students
                elif group.group.program.specialization.qual == 6:
                    superv.hours = settings.kp_kr_3 * superv.students
                else:
                    superv.hours = 0
            else:
                superv.hours = 0
        else:
            superv.hours = 0
    elif superv.supervision_type == 3:
        if "Руководство аспирантом" in course.discipline_settings.discipline.Name:
            if group.group.program.specialization.qual == 8:
                superv.hours = settings.graduate * superv.students
            else:
                superv.hours = 0
        else:
            superv.hours = 0
    elif superv.supervision_type == 4:
        if "Руководство магистрантом" in course.discipline_settings.discipline.Name:
            if group.group.program.specialization.qual == 6:
                superv.hours = settings.master * superv.students
            else:
                superv.hours = 0
        else:
            superv.hours = 0
    elif superv.supervision_type == 5:
        if "Руководство программой магистратуры" in course.discipline_settings.discipline.Name:
            if group.group.program.specialization.qual == 6:
                superv.hours = settings.master_prog
            else:
                superv.hours = 0
        else:
            superv.hours = 0
    else:
        superv.hours = 0
    superv.edu_period = EduPeriod.objects.get(active=True)
    superv.cathedra = cathedra
    superv.save()


def add_practice_hours(teacher, group, cathedra, practice_type, practice):
    settings = HoursSettings.objects.filter(is_active=True).first()
    if practice is None:
        practice = PracticeHours()
    practice.teacher = teacher
    practice.group = group
    practice.practice_type = practice_type
    c = CourseHours.objects.filter(teacher_id=teacher.id, group_id=group.id).first()
    if "практика" in c.discipline_settings.discipline.Name:
        if practice.practice_type == 1:
            if "учебная" in c.discipline_settings.discipline.Name:
                if group.edu_type == 1:
                    if group.group.program.specialization.qual == 4:
                        practice.hours = settings.practice_edu_1 * c.discipline_settings.practice_weekly + settings.practice_preparing
                    elif group.group.program.specialization.qual == 6:
                        practice.hours = settings.practice_edu_2 * group.amount * settings.practice_weekly + settings.practice_preparing
                    else:
                        practice.hours = 0
                elif group.edu_type == 2:
                    if group.group.program.specialization.qual == 4:
                        practice.hours = settings.practice_edu_3 * group.amount + settings.practice_preparing
                    elif group.group.program.specialization.qual == 6:
                        practice.hours = settings.practice_edu_4 * group.amount * settings.practice_weekly + settings.practice_preparing
                    else:
                        practice.hours = 0
                else:
                    practice.hours = 0
            else:
                practice.hours = 0
        elif practice.practice_type == 2:
            if "производственная" in c.discipline_settings.discipline.Name:
                if group.edu_type == 1:
                    if group.group.program.specialization.qual == 4:
                        practice.hours = settings.practice_intern_1 * c.discipline_settings.practice_weekly + settings.practice_preparing
                    elif group.group.program.specialization.qual == 6:
                        practice.hours = settings.practice_intern_2 * group.amount * c.discipline_settings.practice_weekly + settings.practice_preparing
                    else:
                        practice.hours = 0
                elif group.edu_type == 2:
                    if group.group.program.specialization.qual == 4:
                        practice.hours = group.amount * settings.practice_intern_3 + settings.practice_preparing
                    elif group.group.program.specialization.qual == 6:
                        practice.hours = c.discipline_settings.practice_weekly * settings.practice_intern_4 * group.amount + settings.practice_preparing
                    else:
                        practice.hours = 0
                else:
                    practice.hours = 0
            else:
                practice.hours = 0
        elif practice.practice_type == 3:
            if "преддипломная" in c.discipline_settings.discipline.Name:
                if group.group.program.specialization.qual == 4:
                    practice.hours = settings.practice_graduate_1 * c.discipline_settings.practice_weekly + settings.practice_preparing
                elif group.group.program.specialization.qual == 6:
                    practice.hours = settings.practice_graduate_2 * c.discipline_settings.practice_weekly * group.amount + settings.practice_preparing
                else:
                    practice.hours = 0
            else:
                practice.hours = 0
        elif practice.practice_type == 4:
            if "педагогическая" in c.discipline_settings.discipline.Name:
                practice.hours = settings.practice_teaching
            else:
                practice.hours = 0
        else:
            practice.hours = 0
    else:
        practice.hours = 0
    practice.edu_period = EduPeriod.objects.get(active=True)
    practice.cathedra = cathedra
    practice.save()


def add_other_hours(teacher, group, cathedra, other_type, other):
    settings = HoursSettings.objects.get(is_active=True)
    if other is None:
        other = OtherHours()
    other.teacher = teacher
    other.group = group
    other.other_type = other_type
    c = CourseHours.objects.filter(teacher_id=teacher.id, group_id=group.id).first()
    gek = TeacherGekStatus.objects.filter(teacher_id=teacher.id).first()
    if other.other_type == 1:
        if gek:
            if gek.status == 1:
                other.hours = settings.gek_1 * group.amount
            elif gek.status == 2:
                other.hours = settings.gek_2 * group.amount
            elif gek.status == 3:
                other.hours = settings.gek_2 * group.amount
            elif gek.status == 4:
                other.hours = settings.gek_3 * group.amount
            else:
                other.hours = 0
        else:
            other.hours = 0
    elif other.other_type == 2:
        if "Рецензирование магистерских диссертаций" in c.discipline_settings.discipline.Name:
            if group.group.program.specialization.qual == 6:
                other.hours = settings.vkr_rev_1 * group.amount
            elif group.group.program.specialization.qual == 8:
                other.hours = settings.vkr_rev_2 * group.amount
            else:
                other.hours = 0
        else:
            other.hours = 0
    elif other.other_type == 3:
        if group.group.program.specialization.qual == 8:
            other.hours = settings.admis * group.amount
        else:
            other.hours = 0
    elif other.other_type == 4:
        if "Рецензирование рефератов" in c.discipline_settings.discipline.Name:
            other.hours = settings.ref_rev * group.amount
        else:
            other.hours = 0
    else:
        other.hours = 0
    other.edu_period = EduPeriod.objects.get(active=True)
    other.cathedra = cathedra
    other.save()
