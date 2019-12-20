from datetime import *

from django.urls import reverse_lazy, reverse
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_protect
from django.views.generic import ListView, CreateView, UpdateView
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Font, Side
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.contrib.auth.decorators import permission_required, login_required
from umo.models import Discipline, DisciplineDetails, ExamMarks, Group, Semester, Teacher, Person, BRSpoints, Course, \
    GroupList, CheckPoint
from umo.objgens import get_check_points, add_brs, add_exam, add_exam_marks
from disciplines.view_excel import discipline_scores_to_excel
from nomenclature.form import AddSubjectToteacherForm
from umo.models import (Teacher, Group, GroupList, Synch, Year, EduProgram, Student, Discipline, CheckPoint, Control,
                        DisciplineDetails, BRSpoints, EduPeriod, ExamMarks, Exam)
from datetime import datetime

from django.contrib.auth.decorators import permission_required
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.contrib.auth.models import Group as auth_groups
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Protection, Font, Side
import disciplines.view_excel as excel_forms


class DisciplineList(PermissionRequiredMixin, ListView):
    permission_required = 'umo.can_view_scores'
    template_name = 'disc_list.html'
    context_object_name = 'discipline_list'

    def get_queryset(self):
        return Course.objects.all()


@login_required
def list_disc(request, pk):
    teacher = Teacher.objects.get(id=pk)
    disciplines = teacher.course_set.select_related('discipline_detail').filter(is_finished=False)
    form = AddSubjectToteacherForm()
    return render(request, 'disc_list.html', {'discipline_list': disciplines, 'form': form, 'teacher': teacher})


@login_required
def teachers_subjects(request):
    try:
        p = Person.objects.get(user__id=request.user.id)
    except:
        return render(request, 'disc_error.html')
    return HttpResponseRedirect(reverse('disciplines:disciplines', kwargs={'pk': p.pk}))


class DisciplineCreate(PermissionRequiredMixin, CreateView):
    permission_required = 'umo.add_discipline'
    template_name = 'disciplines_form.html'
    success_url = reverse_lazy('disciplines:details_add')
    model = Discipline
    fields = [
        'Name',
        'code',
        'program',
        'lecturer',
    ]


class DisciplineUpdate(PermissionRequiredMixin, UpdateView):
    permission_required = 'umo.change_discipline'
    template_name = 'disciplines_update.html'
    success_url = reverse_lazy('disciplines:disciplines_list')
    model = Discipline
    fields = [
        'Name',
        'code',
        'program',
        'lecturer',
    ]


@login_required
@permission_required('umo.delete_discipline', login_url='login')
def discipline_delete(request):
    if request.method == 'POST':
        discipline_ = Discipline.objects.get(pk=request.POST['discipline'])
        discipline_.delete()
        return HttpResponseRedirect(reverse('disciplines:disciplines_list'))


@login_required
def discipline_detail(request, pk):
    if request.method == 'POST':
        pass
    subject = Course.objects.get(id=pk).discipline_detail.discipline  # Discipline.objects.get(id=pk)
    name = subject.Name
    details = DisciplineDetails.objects.filter(discipline__id=subject.id)
    return render(request, 'disciplines_detail.html', {'form': details, 'name': name})


class DisciplineDetailsList(PermissionRequiredMixin, ListView):
    permission_required = 'umo.change_disciplinedetails'
    template_name = 'disc_details.html'
    context_object_name = 'discipline_details'

    def get_queryset(self):
        return DisciplineDetails.objects.all()


class DetailsCreate(PermissionRequiredMixin, CreateView):
    permission_required = 'umo.add_disciplinedetails'
    template_name = 'details_form.html'
    model = DisciplineDetails
    success_url = reverse_lazy('disciplines:disciplines_list')
    fields = [
        'subject',
        'Credit',
        'Lecture',
        'Practice',
        'Lab',
        'KSR',
        'SRS',
        'semester',
    ]


class DisciplineDetailsUpdate(PermissionRequiredMixin, UpdateView):
    permission_required = 'umo.change_disciplinedetails'
    template_name = 'disciplines_update.html'
    success_url = reverse_lazy('disciplines:disciplines_list')
    model = DisciplineDetails
    fields = [
        'subject',
        'Credit',
        'Lecture',
        'Practice',
        'Lab',
        'KSR',
        'SRS',
        'semester',
    ]


def get_data_for_ekran(request):
    group_id = request.GET.get('group', '')
    semester_id = request.GET.get('semester', '')
    group = get_object_or_404(Group, pk=group_id)
    semester = get_object_or_404(Semester, pk=semester_id)
    grouplist = group.grouplist_set.select_related('student').filter(active=True)
    subjects = group.program.discipline_set.filter(disciplinedetails__semester__name=semester)
    result = {'data': []}
    for s in grouplist:
        m = [s.student.FIO]
        for gl in subjects:
            mark = ExamMarks.objects.filter(student__id=s.student.id, exam__discipline__id=gl.id).first()
            if mark is not None:
                m.append(mark.mark.name)
            else:
                m.append('')
        result['data'].append(m)
    return JsonResponse(result)


def subjects(request):
    group_id = request.GET.get('group', '')
    semester_id = request.GET.get('semester', '')
    result = {'data': []}
    if len(group_id) > 0 and len(semester_id) > 0:
        group = get_object_or_404(Group, pk=group_id)
        semester = get_object_or_404(Semester, pk=semester_id)
        subjects = group.program.discipline_set.filter(disciplinedetails__semester__name=semester)

        for s in subjects:
            result['data'].append(s.Name)
    return JsonResponse(result)


def export_to_excel(request):
    # определяем стили
    font_main = Font(name='Times New Roman',
                     size=12,
                     bold=False,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     color='FF000000',
                     )
    font = Font(name='Calibri',
                size=10,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000',
                )
    font_bold = Font(name='Times New Roman',
                     size=16,
                     bold=True,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     color='FF000000',
                     )
    font_small = Font(name='Times New Roman',
                      size=10,
                      bold=False,
                      italic=False,
                      vertAlign=None,
                      underline='none',
                      strike=False,
                      color='FF000000',
                      )

    fill = PatternFill(fill_type='solid',
                       start_color='c1c1c1',
                       end_color='c2c2c2')

    border = Border(left=Side(border_style='thin',
                              color='FF000000'),
                    right=Side(border_style='thin',
                               color='FF000000'),
                    top=Side(border_style='thin',
                             color='FF000000'),
                    bottom=Side(border_style='thin',
                                color='FF000000'),
                    diagonal=Side(border_style='thin',
                                  color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style='thin',
                                 color='FF000000'),
                    vertical=Side(border_style='thin',
                                  color='FF000000'),
                    horizontal=Side(border_style='thin',
                                    color='FF000000')
                    )
    align_center = Alignment(horizontal='center',
                             vertical='center',
                             text_rotation=0,
                             wrap_text=True,
                             shrink_to_fit=False,
                             indent=0)
    align_right = Alignment(horizontal='right',
                            vertical='center',
                            text_rotation=0,
                            wrap_text=True,
                            shrink_to_fit=False,
                            indent=0)
    align_vertical = Alignment(horizontal='left',
                               vertical='bottom',
                               text_rotation=90,
                               wrap_text=True,
                               shrink_to_fit=False,
                               indent=0)

    # объект
    wb = Workbook()

    # активный лист
    ws = wb.active

    # название страницы
    # ws = wb.create_sheet('первая страница', 0)
    ws.title = 'первая страница'

    # значение ячейки
    # ws['A1'] = "Hello!"

    # текущее время
    today = datetime.today()
    today = today.strftime('%d.%m.%Y %S:%M:%H')

    # данные для строк
    group_id = request.GET['dropdown1']
    semester = request.GET['dropdown2']
    group = Group.objects.get(pk=group_id)
    students = group.grouplist_set.all().order_by('student__FIO')
    subjects = group.program.discipline_set.filter(disciplinedetails__semester__name=semester)
    _row = 3
    _column = 3
    i = 1
    z = 0
    x = 0
    ws.cell(row=1, column=2).value = group.Name + ' cеместр ' + semester
    ws.cell(row=2, column=2).value = 'Всего часов/ЗЕТ'
    for s in subjects:
        ws.cell(row=1, column=_column).value = s.Name
        details = s.disciplinedetails_set.filter(semester__name=semester)
        for k in details:
            ws.cell(row=2, column=_column).value = str(k.total_hours)
        _column += 1
        x += 1
    for gl in students:
        ws.cell(row=_row, column=1).value = str(i) + '.'
        ws.cell(row=_row, column=2).value = gl.student.FIO
        _column = 3
        i += 1
        for s in subjects:
            mark = ExamMarks.objects.filter(student__id=gl.student.id,
                                            exam__course__discipline_detail__discipline__id=s.id).first()
            if mark is not None:
                ws.cell(row=_row, column=_column).value = mark.mark.name
            _column += 1
        _row += 1
        z += 1

    zk = z + 2
    xk = x + 2

    if xk == 2:
        xk = 'B'
    elif xk == 3:
        xk = 'C'
    elif xk == 4:
        xk = 'D'
    elif xk == 5:
        xk = 'E'
    elif xk == 6:
        xk = 'F'
    elif xk == 7:
        xk = 'G'
    elif xk == 8:
        xk = 'H'
    elif xk == 9:
        xk = 'I'
    elif xk == 10:
        xk = 'J'
    elif xk == 11:
        xk = 'K'
    elif xk == 12:
        xk = 'L'
    elif xk == 13:
        xk = 'M'
    elif xk == 14:
        xk = 'N'
    elif xk == 15:
        xk = 'O'
    elif xk == 16:
        xk = 'P'
    elif xk == 17:
        xk = 'Q'
    elif xk == 18:
        xk = 'R'
    elif xk == 19:
        xk = 'S'
    elif xk == 20:
        xk = 'T'

    # шрифты
    for cellObj in ws['A1:' + str(xk) + str(zk)]:
        for cell in cellObj:
            ws[cell.coordinate].font = font_main
    for cellObj in ws['C2:' + str(xk) + '2']:
        for cell in cellObj:
            ws[cell.coordinate].font = font_small
    for cellObj in ws['C1:' + str(xk) + '1']:
        for cell in cellObj:
            ws[cell.coordinate].font = font
    ws['B1'].font = font_bold

    # увеличиваем все строки по высоте
    max_row = ws.max_row
    i = 1
    while i <= max_row:
        rd = ws.row_dimensions[i]
        rd.height = 16
        i += 1

    # вручную устанавливаем высоту первой строки
    rd = ws.row_dimensions[1]
    rd.height = 90

    # сетка
    for cellObj in ws['A1:' + str(xk) + str(zk)]:
        for cell in cellObj:
            # print(cell.coordinate, cell.value)
            ws[cell.coordinate].border = border

    # закрашивание столбца
    for cellObj in ws['A2:' + str(xk) + '2']:
        for cell in cellObj:
            ws[cell.coordinate].fill = fill

    # выравнивание столбца
    for cellObj in ws['A1:' + str(xk) + '1']:
        for cell in cellObj:
            ws[cell.coordinate].alignment = align_vertical

    ws['B1'].alignment = align_center
    ws['B2'].alignment = align_right

    # перетягивание ячеек
    dims = {}
    for cellObj in ws['B1:B' + str(zk)]:
        for cell in cellObj:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
    for col, value in dims.items():
        # value * коэфициент
        ws.column_dimensions[col].width = value * 1.5

    # перетягивание ячеек номеров
    dims = {}
    for cellObj in ws['A1:A' + str(zk)]:
        for cell in cellObj:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
    for col, value in dims.items():
        # value * коэфициент
        ws.column_dimensions[col].width = value * 1.5

    # сохранение файла в выбранную директорию
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=items.xlsx'

    wb.save(response)

    return response


@permission_required('umo.add_discipline', login_url='login')
def excel(request):
    groupname = Group.objects.all().order_by('Name')
    semester_name = Semester.objects.all().order_by('-name')
    subjects = Discipline.objects.all().order_by('Name')

    return render(request, 'export_to_excel.html', {'groupname': groupname, 'semester_name': semester_name,
                                                    'subjects': subjects})


class StudentsScoresView(PermissionRequiredMixin, ListView):
    model = BRSpoints
    permission_required = 'umo.can_view_scores'
    template_name = 'students_scores.html'

    def add_group_brs_points(self):
        pass

    def get_queryset(self):
        #students = course.group.grouplist_set.filter(active=True).values_list('student__id', flat=True)
        return BRSpoints.objects.filter(course__id=self.kwargs['pk'], student__grouplist__active=True).select_related('student', 'checkpoint')

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        course = Course.objects.select_related('discipline_detail').get(pk=self.kwargs['pk'])
        checkpoints = get_check_points()
        if not course.is_finished:
            group_students = course.group.grouplist_set.select_related('student', 'group').filter(active=True).order_by('student__FIO')
            students_to_add = list(set(group_students.values_list('student__id', flat=True)) - set(context['object_list'].values_list('student__id', flat=True)))
            if len(students_to_add) > 0:
                #calc students to add
                add_brs(course, GroupList.objects.filter(student__id__in=students_to_add), checkpoints)
                context['object_list'] = BRSpoints.objects.filter(course__id=self.kwargs['pk'], student__id__in=group_students.values_list('student__id', flat=True)).select_related('student', 'checkpoint')
        else:
            students = set(context['object_list'].values_list('student__id', flat=True))
            group_students = course.group.grouplist_set.select_related('student', 'group').filter(student__id__in=students).order_by('student__FIO')
        #elif len(context['object_list']) // 3
        context['points'] = {}
        context['maxpoints'] = {}
        for item in context['object_list']:
            if item.student.id not in context['points']:
                context['points'][item.student.id] = {}
            context['points'][item.student.id][item.checkpoint.id] = item.points
        for mpoint in course.coursemaxpoints_set.all():
            context['maxpoints'][mpoint.checkpoint.id] = mpoint.max_point
        context['checkpoints'] = checkpoints
        context['group_list'] = group_students
        context['discipline'] = course
        return context


@login_required
@permission_required('umo.can_view_scores', login_url='login')
def export_brs_points(request):
    wb = discipline_scores_to_excel(request.GET['course_id'])
    if wb is not None:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=vedomost.xlsx'
        wb.save(response)
    else:
        response = HttpResponse()
        response.status_code = 404
        response.write('<p>Ошибка при формировании отчета!!</p>')
    return response


class ExamPointsListView(PermissionRequiredMixin, ListView):
    model = GroupList
    context_object_name = 'students_list'
    success_url = reverse_lazy('disciplines:disciplines_list')
    template_name = "exam_points.html"
    permission_required = 'umo.add_exammarks'

    def get_queryset(self):
        course = Course.objects.select_related('discipline_detail', 'group').get(pk=self.kwargs['pk'])
        students = course.group.grouplist_set.values_list('student__id', flat=True)
        return ExamMarks.objects.filter(exam__course__id=self.kwargs['pk'], student__id__in=students).select_related('student', 'mark').order_by('student__FIO')

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        course = Course.objects.select_related('discipline_detail', 'group').get(pk=self.kwargs['pk'])
        group_students = course.group.grouplist_set.select_related('student', 'group').filter(active=True)
        control = course.discipline_detail.control_set.filter(id=self.kwargs['control_id']).first()
        exam = Exam.objects.filter(course__id=self.kwargs['pk'], controlType=control.control_type).first()
        if exam is None:
            exam = add_exam(course, group_students, datetime.today(), control.control_type)
        elif not exam.is_finished:
            with transaction.atomic():
                add_exam_marks(exam, group_students)
        context['object_list'] = ExamMarks.objects.filter(exam__id=exam.id).select_related('student').order_by('student__FIO')
        context['exam'] = exam
        context['group_list'] = group_students
        context['discipline'] = course
        context['today'] = exam.examDate.strftime('%Y-%m-%d')
        context['period'] = EduPeriod.objects.get(active=True)
        context['control_type'] = Control.CONTROL_FORM[control.control_type][1]
        if control.control_type == 1:
            context['is_exam'] = True
        return context


@login_required
@permission_required('umo.can_view_scores', login_url='login')
def exam_report(request):
    wb = excel_forms.exam_scores(int(request.GET['exam_id']))
    if wb is not None:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=vedomost.xlsx'
        wb.save(response)
    else:
        response = HttpResponse()
        response.status_code = 404
        response.write('<p>Ошибка при формировании отчета!!</p>')
    return response