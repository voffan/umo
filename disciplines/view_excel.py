import datetime
import os

from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook, load_workbook
from openpyxl.descriptors import Bool
from openpyxl.styles import PatternFill, Border, Alignment, Protection, Font, Side, NamedStyle
from openpyxl.worksheet.page import PrintPageSetup
from openpyxl.worksheet.properties import WorksheetProperties

from umo.models import Exam, Group, Student, EduPeriod, Course, Course, ExamMarks, GroupList, Control


def get_styles():
    font_main = Font(name='Times New Roman',
                     size=12,
                     bold=False,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     color='FF000000')

    font_bold = Font(name='Times New Roman',
                     size=12,
                     bold=True,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     color='FF000000')

    font_bold_s = Font(name='Times New Roman',
                       size=10,
                       bold=True,
                       italic=False,
                       vertAlign=None,
                       underline='none',
                       strike=False,
                       color='FF000000')

    font_calibri = Font(name='Calibri',
                        size=11,
                        bold=False,
                        italic=False,
                        vertAlign=None,
                        underline='none',
                        strike=False,
                        color='FF000000')

    font_arial = Font(name='Arial Cyr',
                      size=12,
                      bold=False,
                      italic=True,
                      vertAlign=None,
                      underline='none',
                      strike=False,
                      color='FF000000')

    fill = PatternFill(fill_type='solid',
                       start_color='c1c1c1',
                       end_color='c2c2c2')

    border = Border(left=Side(border_style='thin',
                              color='FF000000'),
                    right=Side(border_style='thin',
                               color='FF000000'),
                    top=Side(border_style='thin',
                             color='FF000000'),
                    bottom=Side(border_style='thin',
                                color='FF000000'),
                    diagonal=Side(border_style='thin',
                                  color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style='thin',
                                 color='FF000000'),
                    vertical=Side(border_style='thin',
                                  color='FF000000'),
                    horizontal=Side(border_style='thin',
                                    color='FF000000')
                    )
    align_center = Alignment(horizontal='center',
                             vertical='center',
                             text_rotation=0,
                             wrap_text=False,
                             shrink_to_fit=False,
                             indent=0)
    align_center2 = Alignment(horizontal='center',
                              vertical='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    align_left = Alignment(horizontal='left',
                           vertical='center',
                           text_rotation=0,
                           wrap_text=False,
                           shrink_to_fit=False,
                           indent=0)
    align_right = Alignment(horizontal='right',
                            vertical='center',
                            text_rotation=0,
                            wrap_text=False,
                            shrink_to_fit=False,
                            indent=0)
    header_font = Font(name='Times New Roman',
                             size=9,
                             bold=False,
                             italic=False,
                             vertAlign=None,
                             underline='none',
                             strike=False,
                             color='FF000000')
    return (font_main, font_bold, font_bold_s, font_calibri, font_arial, fill, border, align_center, align_center2, align_left, align_right, header_font)


def discipline_scores_to_excel(course_id):
    # определяем стили
    font_main, font_bold, font_bold_s, font_calibri, font_arial, fill, border, align_center, align_center2, align_left, align_right, header_font = get_styles()
    number_format = 'General'
    protection = Protection(locked=True,
                            hidden=False)

    # объект
    wb = Workbook()
    # активный лист
    ws = wb.active
    # название страницы
    # ws = wb.create_sheet('первая страница', 0)
    ws.title = 'первая страница'

    # текущее время
    p = datetime.datetime.now()

    # данные для строк
    course = Course.objects.select_related('group','discipline_detail','lecturer').get(pk=course_id)
    semester = course.discipline_detail.semester
    group = course.group
    checkpoints = course.coursemaxpoints_set.all()
    if not course or not semester or not group or not checkpoints:
        return None
    edu_period = get_object_or_404(EduPeriod, active=True)


    _row = 13
    _column = 4
    k = 1
    z = 0

    for gl in group.grouplist_set.select_related('student', 'group').filter(active=True).order_by('student__FIO'):
        ws.cell(row=_row, column=1).value = str(k)
        ws.cell(row=_row, column=2).value = gl.student.FIO
        ws.cell(row=_row, column=3).value = gl.student.student_id
        totalpoints = 0
        _column = 4
        for points in gl.student.brspoints_set.filter(course__id=course_id):
            ws.cell(row=_row, column=_column).value = str(points.points).replace('.', ',')
            totalpoints += points.points
            _column += 1

        _row += 1
        k += 1
        z += 1

    zk = z + 11
    zp = z + 14

    if course.lecturer is None:
        fio = 'отсутствует'
    else:
        fio = course.lecturer.FIO

    ws.cell(row=1, column=1).value = 'ФГАОУ ВО «Северо-Восточный федеральный университет им.М.К.Аммосова'
    ws.cell(row=2, column=1).value = 'Институт математики и информатики'
    ws.cell(row=4, column=1).value = 'Контрольный лист текущей и промежуточной аттестации'
    ws.cell(row=5, column=1).value = str(edu_period.begin_year) + '-' + str(edu_period.end_year) + ' учебный год'
    ws.cell(row=6, column=2).value = 'Семестр'
    ws.cell(row=6, column=3).value = semester.name
    ws.cell(row=7, column=2).value = 'Курс'
    ws.cell(row=7, column=3).value = group.year
    ws.cell(row=7, column=5).value = 'Группа'
    ws.cell(row=7, column=6).value = group.Name
    ws.cell(row=8, column=2).value = 'Дисциплина:'
    ws.cell(row=8, column=3).value = course.discipline_detail.discipline.Name
    ws.cell(row=9, column=2).value = 'Преподаватель:'
    ws.cell(row=9, column=3).value = fio
    ws.cell(row=11, column=1).value = '№'
    ws.cell(row=11, column=2).value = 'Фамилия, имя, отчество'
    ws.cell(row=11, column=3).value = '№ зачетной книжки'
    ws.cell(row=11, column=4).value = 'Сумма баллов за текущую работу'
    ws.cell(row=11, column=6).value = 'Рубежный срез'
    ws.cell(row=11, column=7).value = 'Баллы за экзамен(бонусные баллы)'
    ws.cell(row=11, column=8).value = 'Всего баллов'
    ws.cell(row=11, column=9).value = 'Оценка прописью'
    ws.cell(row=11, column=10).value = 'Буквенный эквивалент'
    ws.cell(row=11, column=11).value = 'Подпись преподавателя'
    ws.cell(row=12, column=4).value = '1 контр. срез (max=' + str(int(checkpoints.get(checkpoint__id=4).max_point)) + ')'
    ws.cell(row=12, column=5).value = '2 контр. срез (max=' + str(int(checkpoints.get(checkpoint__id=5).max_point)) + ')'


    # объединение ячеек
    ws.merge_cells('A1:K1')
    ws.merge_cells('A2:K2')
    ws.merge_cells('A4:K4')
    ws.merge_cells('A5:K5')
    ws.merge_cells('A11:A12')
    ws.merge_cells('B11:B12')
    ws.merge_cells('C11:C12')
    ws.merge_cells('D11:E11')
    ws.merge_cells('F11:F12')
    ws.merge_cells('G11:G12')
    ws.merge_cells('H11:H12')
    ws.merge_cells('I11:I12')
    ws.merge_cells('J11:J12')
    ws.merge_cells('K11:K12')

    # шрифты
    font_9 = Font(name='Times New Roman',
                  size=9,
                  bold=False,
                  italic=False,
                  vertAlign=None,
                  underline='none',
                  strike=False,
                  color='FF000000')
    ws['A11'].font = font_9
    ws['B11'].font = font_9
    ws['C11'].font = font_9
    ws['D11'].font = font_9
    ws['D12'].font = font_9
    ws['E12'].font = font_9
    ws['F11'].font = font_9
    ws['G11'].font = font_9
    ws['H11'].font = font_9
    ws['I11'].font = font_9
    ws['J11'].font = font_9
    ws['K11'].font = font_9

    ws['A4'].font = font_bold
    ws['A5'].font = font_bold

    # увеличиваем все строки по высоте
    max_row = ws.max_row
    i = 1
    while i <= max_row:
        rd = ws.row_dimensions[i]
        rd.height = 16
        i += 1

    # вручную устанавливаем высоту первой строки
    rd = ws.row_dimensions[11]
    rd.height = 28
    rd = ws.row_dimensions[12]
    rd.height = 36

    # сетка
    for cellObj in ws['A11:K' + str(zk+1)]:
        for cell in cellObj:
            # print(cell.coordinate, cell.value)
            ws[cell.coordinate].border = border

    # выравнивание

    ws['B6'].alignment = align_right
    ws['B7'].alignment = align_right
    ws['B8'].alignment = align_right
    ws['B9'].alignment = align_right
    ws['E7'].alignment = align_right
    ws['C6'].alignment = align_center
    ws['C7'].alignment = align_center

    for cellObj in ws['A13:A' + str(zk+1)]:
        for cell in cellObj:
            cell.alignment = align_center
    for cellObj in ws['C13:K' + str(zk+1)]:
        for cell in cellObj:
            cell.alignment = align_center

    for cellObj in ws['A1:K5']:
        for cell in cellObj:
            # print(cell.coordinate, cell.value)
            ws[cell.coordinate].alignment = align_center

    for cellObj in ws['A11:K12']:
        for cell in cellObj:
            # print(cell.coordinate, cell.value)
            ws[cell.coordinate].alignment = align_center2

    # перетягивание ячеек
    dims = {}
    for cellObj in ws['A11:A' + str(zk)]:
        for cell in cellObj:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
    for col, value in dims.items():
        # value * коэфициент
        ws.column_dimensions[col].width = value * 3

    dims = {}
    for cellObj in ws['B11:B' + str(zk)]:
        for cell in cellObj:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
    for col, value in dims.items():
        # value * коэфициент
        ws.column_dimensions[col].width = value * 1.2

    return wb


def exam_scores(exam_id):
    # Читаем данные из базы данных
    exam = Exam.objects.select_related('course').get(id=exam_id)
    group = exam.course.group
    exam_points = exam.exammarks_set.all().order_by('student__FIO')
    semester = exam.course.discipline_detail.semester.name
    if int(semester) % 2 == 1:
        additional = int(semester) // 2 
    else:
        additional = (int(semester) - 1) // 2
    edu_period = EduPeriod.objects.get(begin_year__year=group.begin_year.year + additional)

    # Открываем шаблон
    app_dir = os.path.dirname(os.path.abspath(__file__))
    template = os.path.join(app_dir, 'exam_scores.xlsx')
    workbook = load_workbook(template)
    ws = workbook.active

    # Семестр
    ws['A5'] = ws['L5'] = 'Семестр: {}, {} – {} уч.г.'.format(
        semester, 
        edu_period.begin_year.year,
        edu_period.end_year.year,
    )
    
    # Форма контроля
    ws['A6'] = ws['L6'] = 'Форма контроля: {}, курс: {}, группа: {}'.format(
        exam.get_controlType_display(),
        group.year,
        group.Name,
    )

    # Дисциплина
    ws['A7'] = ws['L7'] = 'Дисциплина: ' + exam.course.discipline_detail.discipline.Name

    # ФИО преподавателя
    ws['A8'] = ws['L8'] = 'Фамилия, имя, отчество преподавателя: ' + exam.course.lecturer.FIO

    # Дата
    ws['A9'] = ws['L9'] = 'Дата проведения зачета/экзамена: {:%d.%m.%Y}'.format(exam.examDate)

    # Таблица с баллами
    summary = {ExamMarks.MARKS[0][0]: 0, ExamMarks.MARKS[1][0]: 0, ExamMarks.MARKS[2][0]: 0, ExamMarks.MARKS[3][0]: 0,
               ExamMarks.MARKS[4][0]: 0, ExamMarks.MARKS[5][0]: 0, ExamMarks.MARKS[6][0]: 0, ExamMarks.MARKS[7][0]: 0,
               ExamMarks.MARKS[8][0]: 0, ExamMarks.MARKS[9][0]: 0}
    k = 0
    for points in exam_points:
        k += 1
        ws.insert_rows(11 + k)
        ws.row_dimensions[11 + k].height = 30
        row = str(11 + k)
        ws['A' + row] = ws['L' + row] = str(k)
        ws['B' + row] = ws['M' + row] = points.student.FIO
        ws['C' + row] = ws['N' + row] = points.student.student_id
        ws['D' + row] = ws['O' + row] = points.inPoints + points.additional_points
        ws['E' + row] = ws['P' + row] = points.examPoints
        ws['F' + row] = ws['Q' + row] = points.total_points
        ws['G' + row] = ws['R' + row] = points.get_mark_display()
        ws['H' + row] = ws['S' + row] = points.mark_symbol
        summary[points.mark]+=1

    # Стиль для ячеек таблицы
    solid_line = Side(style='thin', color='000000')
    cell_style = NamedStyle(name='cell_style')
    cell_style.alignment.horizontal = 'center'
    cell_style.alignment.vertical = 'center'
    cell_style.alignment.wrapText = Bool(True)
    cell_style.border = Border(left=solid_line, right=solid_line, top=solid_line, bottom=solid_line)
    cell_style.font = Font(name='Arial', size=9)
    cell_style.number_format = '#.0'

    # Применяем стили к таблице
    for i in range(12, k + 12):
        for j in range(1, 10):
            ws.cell(row=i, column=j).style = ws.cell(row=i, column=j+11).style = cell_style

    ws['C' + str(k + 14)] = ws['N' + str(k + 14)] = summary[6]
    ws['C' + str(k + 15)] = ws['N' + str(k + 15)] = summary[7]
    ws['C' + str(k + 16)] = ws['N' + str(k + 16)] = summary[8]
    ws['C' + str(k + 17)] = ws['N' + str(k + 17)] = summary[5]
    ws['C' + str(k + 18)] = ws['N' + str(k + 18)] = summary[4]
    ws['C' + str(k + 19)] = ws['N' + str(k + 19)] = summary[3]
    ws['C' + str(k + 20)] = ws['N' + str(k + 20)] = summary[2]
    ws['C' + str(k + 21)] = ws['N' + str(k + 21)] = summary[0]
    # Суммы баллов и буквенные эквиваленты оценки
    for i in range(8):
        row = str(14 + k + i)
        ws.merge_cells('E' + row + ':F' + row)
        ws.merge_cells('P' + row + ':Q' + row)
        ws.merge_cells('G' + row + ':I' + row)
        ws.merge_cells('R' + row + ':T' + row)
        ws['E' + row].style = ws['F' + row].style = ws['G' + row].style = \
        ws['H' + row].style = ws['I' + row].style = ws['P' + row].style = \
        ws['Q' + row].style = ws['R' + row].style = ws['S' + row].style = \
        ws['T' + row].style = cell_style

    # Подпись директора
    row = str(24 + k)
    ws.merge_cells('A' + row + ':I' + row)
    ws.merge_cells('L' + row + ':T' + row)

    ws.print_area = 'A1:T' + str(row)
    ws.page_setup = PrintPageSetup(worksheet=ws)
    ws.page_setup.paperSize = '9'
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToHeight = True
    ws.page_setup.fitToWidth = False
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    return workbook
