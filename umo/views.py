from datetime import datetime

from django.contrib.auth.decorators import permission_required, login_required
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.contrib.auth.models import Group as auth_groups, User
from django.contrib.auth.password_validation import validate_password
from django.contrib.auth import login, update_session_auth_hash
from django.core.validators import validate_email
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.forms import ModelForm, CharField, ValidationError, PasswordInput, HiddenInput
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Protection, Font, Side
import csv
import synch.models as sync_models
from umo.models import (Teacher, Group, GroupList, Synch, Year, EduProgram, Student, Discipline, CheckPoint, Control,
                        DisciplineDetails, BRSpoints, EduPeriod, ExamMarks, Exam, Kafedra,Position)
from umo.forms import UploadUsersForm
from nomenclature.views import hadle_uploaded_file


def index(request):
    if auth_groups.objects.get(name='teacher') in request.user.groups.all():
        return redirect('disciplines:mysubjects')
    else:
        return redirect('disciplines:disciplines_list')


class TeacherList(PermissionRequiredMixin, ListView):
    permission_required = 'umo.add_teacher'
    template_name = 'teachers_list.html'
    model = Teacher


class TeacherCreate(PermissionRequiredMixin, CreateView):
    permission_required = 'umo.change_teacher'
    template_name = 'teacher_form.html'
    success_url = reverse_lazy('teachers:list_teachers')
    model = Teacher
    fields = ['FIO', 'position', 'cathedra', 'user']


class TeacherUpdate(PermissionRequiredMixin, UpdateView):
    permission_required = 'umo.change_teacher'
    template_name = 'teacher_edit.html'
    success_url = reverse_lazy('teachers:list_teachers')
    model = Teacher
    fields = ['last_name', 'first_name', 'second_name', 'position', 'title', 'cathedra']


class TeacherDelete(PermissionRequiredMixin, DeleteView):
    permission_required = 'umo.delete_teacher'
    model = Teacher
    success_url = reverse_lazy('teachers:list_teachers')
    template_name = 'teacher_delete.html'


class GroupsList(PermissionRequiredMixin, ListView):
    permission_required = 'umo.add_group'
    template_name = 'groups_list.html'
    model = Group


class TeacherProfileForm(ModelForm):
    #success_url = reverse_lazy('teachers:list_teachers')
    email = CharField(max_length=50, label='Email', required=True)
    current_password = CharField(max_length=100, label='Текущий пароль', required=False, widget=PasswordInput)
    password = CharField(max_length=100, label='Пароль', required=False, widget=PasswordInput)
    confirmation = CharField(max_length=100, label='Подтверждение пароля', required=False, widget=PasswordInput)

    def __init__(self, *args, **kwargs):
        super(ModelForm, self).__init__(*args, **kwargs)
        self.fields['position'].widget.attrs['readonly'] = True
        self.fields['title'].widget.attrs['readonly'] = True
        self.fields['cathedra'].widget.attrs['readonly'] = True

    class Meta:
        model = Teacher
        fields = ['last_name', 'first_name', 'second_name', 'position', 'title', 'cathedra']
        #widgets = {'id': HiddenInput()}

    def clean(self):
        #cleaned_data = super.clean()
        validate_email(self.cleaned_data['email'])
        crnt_pwd = self.cleaned_data['current_password']
        pwd = self.cleaned_data['password']
        confirmation = self.cleaned_data['confirmation']
        if pwd != confirmation:
            raise ValidationError('Введенные пароли не совпадают!')
        if pwd != '':
            if not self.instance.user.check_password(crnt_pwd):
                raise ValidationError('Текущий пароль неверен!')
            validate_password(pwd)
        return self.cleaned_data

    def save(self, commit=True):
        with transaction.atomic():
            teacher = super(TeacherProfileForm, self).save(commit=False)
            teacher.FIO = self.cleaned_data['last_name'] + ' ' + self.cleaned_data['first_name'] + ' ' + self.cleaned_data['second_name']
            teacher.save()
            user = teacher.user
            if self.cleaned_data['email'] != self.instance.user.email:
                user.email = self.cleaned_data['email']
            if user is not None and self.cleaned_data['password'] != '':
                user.set_password(self.cleaned_data['password'])
            user.save()
            return teacher


@permission_required('umo.change_teacher', login_url='/auth/login')
@login_required
def teacher_profile(request):
    teacher = Teacher.objects.filter(user__id=request.user.id).first()
    success_message = None
    if teacher is not None:
        if request.method == 'POST':
            form = TeacherProfileForm(request.POST, instance=teacher)
            if form.is_valid():
                try:
                    form.save()
                    update_session_auth_hash(request, teacher.user)
                    success_message = 'Профиль успешно сохранен!!'
                except:
                    form.add_error('first_name', 'Ошибка сохранения данных!')
        else:
            form = TeacherProfileForm(instance=teacher, initial={'email': request.user.email})
        return render(request, 'teacher_edit.html', {'form': form, 'success_message': success_message, 'profile': 1})
    return HttpResponse('You are not teacher!!')# render(request, 'teacher_edit.html')
'''
@permission_required('umo.add_teacher', login_url='/auth/login')
def list_teachers(request):
    all = Teacher.objects.all()
    return render(request, 'teachers_list.html', {'teachers': all})
'''

def get_mark(str, value):
    if (str.lower() == 'зачет' or str.lower() == 'зачёт'):
        if (value >= 60):
            return 'зач'
        else:
            return 'нз'
    else:
        if (value >= 85):
            return 'отл'
        elif (value >= 65):
            return 'хор'
        elif (value >= 55):
            return 'удовл'
        else:
            return 'неуд'


def get_mark_vedomost(str, inPoints, examPoints):
    value = inPoints + examPoints
    if (str.lower() == 'зачет' or str.lower() == 'зачёт'):
        if (value >= 60):
            return 'Зачтено'
        else:
            return 'Не зачтено'
    else:
        if (inPoints < 45):
            return 'Не допущен'
        if (value >= 85):
            return 'Отлично'
        elif (value >= 65):
            return 'Хорошо'
        elif (value >= 55):
            return 'Удовлетворительно'
        else:
            return 'Неудовлетворительно'


def get_markSymbol(str, value):
    if (str.lower() == 'зачет' or str.lower() == 'зачёт'):
        return None
    else:
        if (value >= 95):
            return 'A'
        elif (value >= 85):
            return 'B'
        elif (value >= 75):
            return 'C'
        elif (value >= 65):
            return 'D'
        elif (value >= 55):
            return 'E'
        elif (value >= 25):
            return 'FX'
        else:
            return 'F'


class BRSPointsListView(ListView):
    model = GroupList
    context_object_name = 'students_list'
    success_url = reverse_lazy('disciplines:disciplines_list')
    template_name = "brs_students.html"

    def get_queryset(self):
        disc = Discipline.objects.get(id = self.kwargs['pk'])
        return GroupList.objects.filter(group__program = disc.program).filter(active=True)

    def get_context_data(self, **kwargs):
        context = super(BRSPointsListView, self).get_context_data(**kwargs)
        checkpoint = CheckPoint.objects.all()
        if (checkpoint.count() < 5):
            if (checkpoint.count() > 0):
                i = 0
                for ch in checkpoint:
                    i = i + 1
                    if (i == 1):
                        ch.Name = "Первый срез"
                    elif (i == 2):
                        ch.Name = "Второй срез"
                    elif (i == 3):
                        ch.Name = "Рубежный срез"
                    else:
                        ch.Name = "Рубежный срез с премиальными баллами"
                    ch.save()
            for i in range(checkpoint.count(), 5):
                ch = CheckPoint()
                if (i == 1):
                    ch.Name = "Первый срез"
                elif (i == 2):
                    ch.Name = "Второй срез"
                elif (i == 3):
                    ch.Name = "Рубежный срез"
                elif (i == 4):
                    ch.Name = "Рубежный срез с премиальными баллами"
                else:
                    ch.Name = "Всего баллов"
                ch.save()
        context['checkpoint'] = checkpoint
        discipline = Discipline.objects.get(id=self.kwargs['pk'])
        control = Control.objects.filter(discipline_detail__subject__id=discipline.id).first()
        context['control_type'] = 'Баллы ' + control.controltype.name.lower()
        context['discipline'] = discipline
        discipline_details = DisciplineDetails.objects.filter(subject=discipline)
        semester = []
        for d in discipline_details:
            semester.append(d.semester)
        context['semester'] = semester
        grouplist = GroupList.objects.filter(group__program=discipline.program).filter(active=True)
        context['grouplist'] = grouplist
        dict = {}
        for s in semester:
            dict[s.name] = {}
            dict[s.name]['key'] = s.name
            control = Control.objects.filter(discipline_detail__subject__id=discipline.id, discipline_detail__semester=s).first()
            for gl in grouplist:
                dict[s.name][str(gl.id)] = {}
                dict[s.name][str(gl.id)]['key'] = gl.id
                i = 0
                for ch in checkpoint:
                    i = i + 1
                    newBRSpoints = BRSpoints.objects.filter(brs__discipline__id = discipline.id, brs__semester=s).filter(CheckPoint = ch).filter(student = gl.student).first()
                    if (newBRSpoints is None):
                        newBRSpoints = BRSpoints()
                        newBRSpoints.student = gl.student
                        newBRSpoints.checkpoint = ch
                        newBRSpoints.points = 0.0
                        newBRS = BRS.objects.filter(discipline__id = discipline.id, semester=s).first()
                        if (newBRS is None):
                            newBRS = BRS()
                            newBRS.discipline = discipline
                            newBRS.eduperiod = EduPeriod.objects.all().first()
                            newBRS.semester = s
                            newBRS.save()
                        newBRSpoints.brs = newBRS
                        newBRSpoints.save()
                    dict[s.name][str(gl.id)][str(i)] = newBRSpoints

                newExamMarks = ExamMarks.objects.filter(exam__discipline__id = discipline.id, exam__semester=s).filter(student = gl.student).first()
                if (newExamMarks is None):
                    # newMarkSymbol = MarkSymbol.objects.filter(name='F').first()
                    # if (newMarkSymbol is None):
                    #     newMarkSymbol = MarkSymbol.objects.create(name='F')

                    newMark = 2

                    newExam = Exam.objects.filter(discipline__id = discipline.id, semester=s).first()
                    if (newExam is None):
                        newExam = Exam()
                        newExam.controlType = control.controltype
                        newExam.course = discipline
                        newExam.eduperiod = EduPeriod.objects.all().first()
                        newExam.examDate = 'не проставлена'
                        newExam.semester = s
                        newExam.save()

                    newExamMarks = ExamMarks()
                    newExamMarks.student = gl.student
                    newExamMarks.inPoints = 0.0
                    newExamMarks.examPoints = 0.0
                    # newExamMarks.markSymbol = newMarkSymbol
                    newExamMarks.mark_symbol = 'F'
                    newExamMarks.mark = newMark
                    newExamMarks.exam = newExam
                    newExamMarks.save()
                dict[s.name][str(gl.id)]['6'] = newExamMarks
        context['dict'] = dict
        return context

    def post(self, request, *args, **kwargs):
        if (request.POST.get('save')):
            studid = request.POST.getlist('studid')
            points = []
            points.append(request.POST.getlist('points1'))
            points.append(request.POST.getlist('points2'))
            points.append(request.POST.getlist('points3'))
            points.append(request.POST.getlist('points4'))
            points.append(request.POST.getlist('points6'))
            semester = request.POST.getlist('semester')
            arr_size = len(studid)
            checkpoint = CheckPoint.objects.all()
            discipline = Discipline.objects.get(id=self.kwargs['pk'])
            exam = Exam.objects.get(discipline__id=self.kwargs['pk'], semester__id=semester[0])
            for i in range(0, arr_size):
                st = Student.objects.get(id=studid[i])
                k = 0

                exammarks = ExamMarks.objects.filter(exam__discipline__id=discipline.id, exam__semester__id=semester[i]).get(student=st)
                exammarks.examPoints = float(points[4][i].replace(',', '.'))
                exammarks.inPoints = float(points[3][i].replace(',', '.'))

                totalPoints = exammarks.examPoints + exammarks.inPoints

                for ch in checkpoint:
                    brspoints = BRSpoints.objects.filter(brs__discipline__id=discipline.id, brs__semester__id=semester[i]).filter(CheckPoint=ch).get(
                        student=st)
                    if (k != 4):
                        brspoints.points = float(points[k][i].replace(',', '.'))
                        k = k + 1
                    else:
                        brspoints.points = totalPoints
                        k = 0
                    brspoints.save()

                tempMarkSymbol = get_markSymbol(exam.controlType.name, totalPoints)
                tempMark = get_mark(exam.controlType.name, totalPoints)

                # if (tempMarkSymbol is None):
                #     newMarkSymbol = None
                # else:
                #     newMarkSymbol = MarkSymbol.objects.filter(name=tempMarkSymbol).first()
                #     if (newMarkSymbol is None):
                #         newMarkSymbol = MarkSymbol.objects.create(name=tempMarkSymbol)

                newMark = 2

                # exammarks.markSymbol = newMarkSymbol
                exammarks.mark_symbol = tempMarkSymbol if tempMarkSymbol else ''
                exammarks.mark = newMark
                exammarks.save()
            return redirect('brs_studentlist', pk=self.kwargs['pk'])

        elif request.POST.get('vedomost'):
            # определяем стили
            font_main = Font(name='Times New Roman',
                             size=12,
                             bold=False,
                             italic=False,
                             vertAlign=None,
                             underline='none',
                             strike=False,
                             color='FF000000')

            font_bold = Font(name='Times New Roman',
                             size=12,
                             bold=True,
                             italic=False,
                             vertAlign=None,
                             underline='none',
                             strike=False,
                             color='FF000000')

            font_bold_s = Font(name='Times New Roman',
                               size=10,
                               bold=True,
                               italic=False,
                               vertAlign=None,
                               underline='none',
                               strike=False,
                               color='FF000000')

            font_calibri = Font(name='Calibri',
                                size=11,
                                bold=False,
                                italic=False,
                                vertAlign=None,
                                underline='none',
                                strike=False,
                                color='FF000000')

            font_arial = Font(name='Arial Cyr',
                              size=12,
                              bold=False,
                              italic=True,
                              vertAlign=None,
                              underline='none',
                              strike=False,
                              color='FF000000')

            fill = PatternFill(fill_type='solid',
                               start_color='c1c1c1',
                               end_color='c2c2c2')

            border = Border(left=Side(border_style='thin',
                                      color='FF000000'),
                            right=Side(border_style='thin',
                                       color='FF000000'),
                            top=Side(border_style='thin',
                                     color='FF000000'),
                            bottom=Side(border_style='thin',
                                        color='FF000000'),
                            diagonal=Side(border_style='thin',
                                          color='FF000000'),
                            diagonal_direction=0,
                            outline=Side(border_style='thin',
                                         color='FF000000'),
                            vertical=Side(border_style='thin',
                                          color='FF000000'),
                            horizontal=Side(border_style='thin',
                                            color='FF000000')
                            )
            align_center = Alignment(horizontal='center',
                                     vertical='center',
                                     text_rotation=0,
                                     wrap_text=False,
                                     shrink_to_fit=False,
                                     indent=0)
            align_center2 = Alignment(horizontal='center',
                                      vertical='center',
                                      text_rotation=0,
                                      wrap_text=True,
                                      shrink_to_fit=False,
                                      indent=0)
            align_left = Alignment(horizontal='left',
                                   vertical='center',
                                   text_rotation=0,
                                   wrap_text=False,
                                   shrink_to_fit=False,
                                   indent=0)
            number_format = 'General'
            protection = Protection(locked=True,
                                    hidden=False)

            # объект
            wb = Workbook()

            # активный лист
            ws = wb.active

            # название страницы
            # ws = wb.create_sheet('первая страница', 0)
            ws.title = 'первая страница'

            # текущее время
            p = datetime.now()

            # данные для строк
            group_name = str(request.POST.get('selected_group'))
            disc_id = self.kwargs['pk']
            studid = request.POST.getlist('studid')
            group = Group.objects.get(Name=group_name)
            inpoints = request.POST.getlist('points4')
            exampoints = request.POST.getlist('points6')
            semester = request.POST.getlist('semester')
            exam = Exam.objects.get(discipline__id=disc_id, semester_id=semester[0])
            arr_size = len(studid)

            _row = 12
            _column = 4
            k = 1
            z = 0

            zachteno = 0
            ne_zachteno = 0
            ne_attest = 0
            otl = 0
            horosho = 0
            udovl = 0
            neudovl = 0
            ne_yavka = 0

            for i in range(0, arr_size):
                gl = Student.objects.get(id=studid[i])
                ws.cell(row=_row, column=1).value = str(k)
                k += 1
                ws.cell(row=_row, column=2).value = gl.FIO
                ws.cell(row=_row, column=3).value = gl.StudentID
                ws.cell(row=_row, column=_column).value = str(float(inpoints[i].replace(',', '.'))).replace('.', ',')
                ws.cell(row=_row, column=_column + 1).value = str(float(exampoints[i].replace(',', '.'))).replace('.', ',')
                totalpoints = float(inpoints[i].replace(',', '.')) + float(exampoints[i].replace(',', '.'))
                ws.cell(row=_row, column=_column + 2).value = str(totalpoints).replace('.', ',')
                ws.cell(row=_row, column=_column + 3).value = get_mark_vedomost(exam.controlType.name, float(inpoints[i].replace(',', '.')), float(exampoints[i].replace(',', '.')))
                ws.cell(row=_row, column=_column + 4).value = get_markSymbol(exam.controlType.name, totalpoints)
                _row += 1
                z += 1

            zk = z+11
            zp = z+14
            zp2 = zp+7

            if group.year is None:
               year = 'отсутствует'
            else:
                year = group.year

            if exam.discipline.lecturer is None:
               fio = 'отсутствует'
            else:
                fio = exam.discipline.lecturer.FIO

            ws.cell(row=1, column=1).value = 'ФГАОУ ВО «Северо-Восточный федеральный университет им.М.К.Аммосова'
            ws.cell(row=2, column=1).value = 'Институт математики и информатики'
            ws.cell(row=3, column=1).value = 'Ведомость текущей и промежуточной аттестации'
            ws.cell(row=5, column=1).value = 'Семестр: ' + str(
                exam.semester.name) + ', ' + exam.eduperiod.beginyear + '-' + exam.eduperiod.endyear + ' уч.г.'
            ws.cell(row=6, column=1).value = 'Форма контроля:'
            ws.cell(row=6, column=3).value = exam.controlType.name
            ws.cell(row=6, column=5).value = 'курс '+str(year)
            ws.cell(row=6, column=6).value = 'группа:'
            ws.cell(row=6, column=7).value = group_name
            ws.cell(row=7, column=1).value = 'Дисциплина:'
            ws.cell(row=7, column=3).value = exam.discipline.Name
            ws.cell(row=8, column=1).value = 'Фамилия, имя, отчество преподавателя:'
            ws.cell(row=8, column=4).value = fio
            ws.cell(row=9, column=1).value = 'Дата проведения зачета/экзамена:'
            ws.cell(row=9, column=3).value = exam.examDate
            ws.cell(row=11, column=1).value = '№'
            ws.cell(row=11, column=2).value = 'Фамилия, имя, отчество'
            ws.cell(row=11, column=3).value = '№ зачетной книжки'
            ws.cell(row=11, column=4).value = 'Сумма баллов за текущую работу-рубеж.срез'
            ws.cell(row=11, column=5).value = 'Баллы ' + exam.controlType.name + ' (бонусные баллы)'
            ws.cell(row=11, column=6).value = 'Всего баллов'
            ws.cell(row=11, column=7).value = 'Оценка прописью'
            ws.cell(row=11, column=8).value = 'Буквенный эквивалент'
            ws.cell(row=11, column=9).value = 'Подпись преподавателя'
            ws.cell(row=zp, column=2).value = 'зачтено'
            ws.cell(row=zp+1, column=2).value = 'не зачтено'
            ws.cell(row=zp+2, column=2).value = 'не аттест'
            ws.cell(row=zp+3, column=2).value = '5(отлично)'
            ws.cell(row=zp+4, column=2).value = '4(хорошо)'
            ws.cell(row=zp+5, column=2).value = '3(удовл)'
            ws.cell(row=zp+6, column=2).value = '2(неудовл)'
            ws.cell(row=zp2, column=2).value = 'не явка'
            ws.cell(row=zp, column=5).value = 'Сумма баллов'
            ws.cell(row=zp+1, column=5).value = '95-100'
            ws.cell(row=zp+2, column=5).value = '85-94,9'
            ws.cell(row=zp+3, column=5).value = '75-84,9'
            ws.cell(row=zp+4, column=5).value = '65-74,9'
            ws.cell(row=zp+5, column=5).value = '55-64,9'
            ws.cell(row=zp+6, column=5).value = '25-54,9'
            ws.cell(row=zp2, column=5).value = '0-24,9'
            ws.cell(row=zp, column=7).value = 'Буквенный эквивалент оценки'
            ws.cell(row=zp+1, column=7).value = 'A'
            ws.cell(row=zp+2, column=7).value = 'B'
            ws.cell(row=zp+3, column=7).value = 'C'
            ws.cell(row=zp+4, column=7).value = 'D'
            ws.cell(row=zp+5, column=7).value = 'E'
            ws.cell(row=zp+6, column=7).value = 'FX'
            ws.cell(row=zp2, column=7).value = 'F'
            ws.cell(row=zp+10, column=2).value = 'Директор ИМИ СВФУ____________________'
            ws.cell(row=zp+10, column=4).value = 'В.И.Афанасьева'

            # объединение ячеек
            ws.merge_cells('A1:I1')
            ws.merge_cells('A2:I2')
            ws.merge_cells('A3:I3')
            ws.merge_cells('A5:B5')
            ws.merge_cells('A6:B6')
            ws.merge_cells('C6:D6')
            ws.merge_cells('A7:B7')
            ws.merge_cells('C7:G7')
            ws.merge_cells('A8:C8')
            ws.merge_cells('D8:G8')
            ws.merge_cells('A9:B9')
            ws.merge_cells('C9:D9')
            ws.merge_cells('E' + str(zp) + ':F' + str(zp))
            ws.merge_cells('E' + str(zp + 1) + ':F' + str(zp + 1))
            ws.merge_cells('E' + str(zp + 2) + ':F' + str(zp + 2))
            ws.merge_cells('E' + str(zp + 3) + ':F' + str(zp + 3))
            ws.merge_cells('E' + str(zp + 4) + ':F' + str(zp + 4))
            ws.merge_cells('E' + str(zp + 5) + ':F' + str(zp + 5))
            ws.merge_cells('E' + str(zp + 6) + ':F' + str(zp + 6))
            ws.merge_cells('E' + str(zp + 7) + ':F' + str(zp + 7))
            ws.merge_cells('G' + str(zp) + ':H' + str(zp))
            ws.merge_cells('G' + str(zp + 1) + ':H' + str(zp + 1))
            ws.merge_cells('G' + str(zp + 2) + ':H' + str(zp + 2))
            ws.merge_cells('G' + str(zp + 3) + ':H' + str(zp + 3))
            ws.merge_cells('G' + str(zp + 4) + ':H' + str(zp + 4))
            ws.merge_cells('G' + str(zp + 5) + ':H' + str(zp + 5))
            ws.merge_cells('G' + str(zp + 6) + ':H' + str(zp + 6))
            ws.merge_cells('G' + str(zp + 7) + ':H' + str(zp + 7))
            ws.merge_cells('B' + str(zp + 10) + ':C' + str(zp + 10))
            ws.merge_cells('D' + str(zp + 10) + ':E' + str(zp + 10))

            for cellObj in ws['G12:G'+str(zk)]:
                for cell in cellObj:
                    if ws[cell.coordinate].value == 'Зачтено':
                        zachteno = zachteno + 1
                    elif ws[cell.coordinate].value == 'Не зачтено':
                        ne_zachteno = ne_zachteno + 1
                    elif ws[cell.coordinate].value == 'Не допущен':
                        ne_attest = ne_attest + 1
                    elif ws[cell.coordinate].value == 'Отлично':
                        otl = otl + 1
                    elif ws[cell.coordinate].value == 'Хорошо':
                        horosho = horosho + 1
                    elif ws[cell.coordinate].value == 'Удовлетворительно':
                        udovl = udovl + 1
                    elif ws[cell.coordinate].value == 'Неудовлетворительно':
                        neudovl = neudovl + 1
                    elif ws[cell.coordinate].value == 'Не явка':
                        ne_yavka = ne_yavka + 1

            ws.cell(row=zp, column=3).value = str(zachteno)
            ws.cell(row=zp+1, column=3).value = str(ne_zachteno)
            ws.cell(row=zp+2, column=3).value = str(ne_attest)
            ws.cell(row=zp+3, column=3).value = str(otl)
            ws.cell(row=zp+4, column=3).value = str(horosho)
            ws.cell(row=zp+5, column=3).value = str(udovl)
            ws.cell(row=zp+6, column=3).value = str(neudovl)
            ws.cell(row=zp2, column=3).value = str(ne_yavka)

            # шрифты
            for cellObj in ws['A1:I'+str(zk)]:
                for cell in cellObj:
                    ws[cell.coordinate].font = font_main

            for cellObj in ws['G12:G'+str(zk)]:
                for cell in cellObj:
                    ws[cell.coordinate].font = font_bold_s

            for cellObj in ws['B12:B'+str(zk)]:
                for cell in cellObj:
                    ws[cell.coordinate].font = font_calibri

            for cellObj in ws['H12:H'+str(zk)]:
                for cell in cellObj:
                    ws[cell.coordinate].font = font_calibri

            for cellObj in ws['E12:E'+str(zk)]:
                for cell in cellObj:
                    ws[cell.coordinate].font = font_bold

            for cellObj in ws['E11:I11']:
                for cell in cellObj:
                    ws[cell.coordinate].font = Font(name='Times New Roman',
                                                  size=9,
                                                  bold=False,
                                                  italic=False,
                                                  vertAlign=None,
                                                  underline='none',
                                                  strike=False,
                                                  color='FF000000')

            ws['A3'].font = font_bold
            ws['C7'].font = font_bold
            ws['D8'].font = font_bold
            ws['F6'].font = font_bold
            ws['C7'].font = font_arial
            ws['D8'].font = font_arial
            ws['G6'].font = Font(name='Arial Cyr',
                                 size=12,
                                 bold=False,
                                 italic=True,
                                 vertAlign=None,
                                 underline='single',
                                 strike=False,
                                 color='FF000000')
            ws['C9'].font = Font(name='Calibri',
                                 size=11,
                                 bold=False,
                                 italic=False,
                                 vertAlign=None,
                                 underline='single',
                                 strike=False,
                                 color='FF000000')
            ws['A11'].font = Font(name='Times New Roman',
                                  size=10,
                                  bold=False,
                                  italic=False,
                                  vertAlign=None,
                                  underline='none',
                                  strike=False,
                                  color='FF000000')
            ws['B11'].font = Font(name='Times New Roman',
                                  size=10,
                                  bold=False,
                                  italic=False,
                                  vertAlign=None,
                                  underline='none',
                                  strike=False,
                                  color='FF000000')
            ws['C11'].font = Font(name='Times New Roman',
                                  size=9,
                                  bold=False,
                                  italic=False,
                                  vertAlign=None,
                                  underline='none',
                                  strike=False,
                                  color='FF000000')
            ws['D11'].font = Font(name='Times New Roman',
                                  size=8,
                                  bold=False,
                                  italic=False,
                                  vertAlign=None,
                                  underline='none',
                                  strike=False,
                                  color='FF000000')
            ws['C6'].font = Font(name='Times New Roman',
                                 size=14,
                                 bold=False,
                                 italic=True,
                                 vertAlign=None,
                                 underline='single',
                                 strike=False,
                                 color='FF000000')

            # увеличиваем все строки по высоте
            max_row = ws.max_row
            i = 1
            while i <= max_row:
                rd = ws.row_dimensions[i]
                rd.height = 16
                i += 1

            # вручную устанавливаем высоту первой строки
            rd = ws.row_dimensions[11]
            rd.height = 48

            # сетка
            for cellObj in ws['A11:I'+str(zk)]:
                for cell in cellObj:
                    # print(cell.coordinate, cell.value)
                    ws[cell.coordinate].border = border

            for cellObj in ws['B'+str(zp)+':C'+str(zp2)]:
                for cell in cellObj:
                    # print(cell.coordinate, cell.value)
                    ws[cell.coordinate].border = border

            for cellObj in ws['E'+str(zp)+':H'+str(zp2)]:
                for cell in cellObj:
                    # print(cell.coordinate, cell.value)
                    ws[cell.coordinate].border = border

            # выравнивание
            for cellObj in ws['A1:I3'+str(zk)]:
                for cell in cellObj:
                    # print(cell.coordinate, cell.value)
                    ws[cell.coordinate].alignment = align_center

            for cellObj in ws['A11:I11']:
                for cell in cellObj:
                    # print(cell.coordinate, cell.value)
                    ws[cell.coordinate].alignment = align_center2

            for cellObj in ws['A5:I9']:
                for cell in cellObj:
                    # print(cell.coordinate, cell.value)
                    ws[cell.coordinate].alignment = align_left

            # перетягивание ячеек
            dims = {}
            for cellObj in ws['G11:G'+str(zk)]:
                for cell in cellObj:
                    if cell.value:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
            for col, value in dims.items():
                # value * коэфициент
                ws.column_dimensions[col].width = value * 1.5

            dims = {}
            for cellObj in ws['A11:A'+str(zk)]:
                for cell in cellObj:
                    if cell.value:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
            for col, value in dims.items():
                # value * коэфициент
                ws.column_dimensions[col].width = value * 3

            dims = {}
            for cellObj in ws['B11:B'+str(zk)]:
                for cell in cellObj:
                    if cell.value:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
            for col, value in dims.items():
                # value * коэфициент
                ws.column_dimensions[col].width = value * 1.5

            dims = {}
            for cellObj in ws['D11:D'+str(zk)]:
                for cell in cellObj:
                    if cell.value:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
            for col, value in dims.items():
                # value * коэфициент
                ws.column_dimensions[col].width = value * 0.25

            # сохранение файла в выбранную директорию
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=vedomost.xlsx'

            wb.save(response)

            return response
        



def add_users(request):
    f=UploadUsersForm()
    if request.method == "POST":
        f = UploadUsersForm(request.POST, request.FILES)
        if f.is_valid():
            file_path = hadle_uploaded_file(request.FILES['file'].name, request.FILES['file'])
            logs=[]
            print(file_path)
            process_users(file_path, logs)
            return render(request,'logs.html')

    return render(request, 'users_upload.html', {'form': f})


def process_users(file_name, logs):
    with open(file_name,newline='') as read:
        #try:
            reader=csv.DictReader(read,delimiter=";")
        #except:
        #try:
            for row in reader:
                u = User()
                u.username = login_gen(row['teacher_last-name'], row['teacher_first-name'], row['teacher_second-name'])
                u.set_password("jkjlkj")
                u.email = "danilove00@mail.ru"
                u.save()
                t = Teacher()
                t.user = u
                #t.save()
                t.FIO = row['teacher_last-name'] + " " + row['teacher_first-name'] + " " + row['teacher_second-name']
                t.last_name = row['teacher_last-name']
                t.first_name = row['teacher_first-name']
                t.second_name = row['teacher_second-name']
                t.maiden_name = row['teacher_maiden-name']
                c_id = int(row['teacher_kafedra'])
                t.cathedra = Kafedra.objects.get(number=c_id)
                t_id = int(row['teacher_title'])
                t.title = t_id

                t.position = Position.objects.get(name = str(row['teacher_position']))
                t.save()
        # except:
        #     print("Ошибка чего-то там")
        
            


def login_gen(surname,name,otch): #сгенерировать логин по ФИО
    return surname+"."+name[0]+"ролрлор"+otch[0]

def pass_gen(): #рандомно сгенерировать пароль
    pass
    
