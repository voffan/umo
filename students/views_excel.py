import datetime

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Protection, Font, Side
from openpyxl.utils.cell import get_column_interval
from openpyxl.worksheet.page import PrintPageSetup

from umo.models import EduPeriod, CourseMaxPoints, CheckPoint, Course, BRSpoints, GroupList


def print_headers(work_sheet, group, semester, courses, start):
    period = EduPeriod.objects.get(active=True)
    work_sheet.cell(row=start, column=1).value = 'ФГАОУ ВО СЕВЕРО-ВОСТОЧНЫЙ ФЕДЕРАЛЬНЫЙ УНИВЕРСИТЕТ им. М.К. АММОСОВА'
    work_sheet.cell(row=start + 1, column=1).value = 'ИНСТИТУТ МАТЕМАТИКИ И ИНФОРМАТИКИ'
    work_sheet.cell(row=start + 2, column=1).value = 'ЖУРНАЛ ТЕКУЩЕЙ АТТЕСТАЦИИ за ' + str(period) + ' уч. год'
    work_sheet.cell(row=start + 4, column=1).value = 'Семестр'
    work_sheet.cell(row=start + 4, column=3).value = semester.name
    work_sheet.cell(row=start + 5, column=1).value = 'Курс'
    work_sheet.cell(row=start + 6, column=1).value = 'Направление подготовки'
    work_sheet.cell(row=start + 5, column=3).value = group.year
    work_sheet.cell(row=start + 6, column=3).value = group.program.specialization.code + ' ' + group.program.specialization.name
    work_sheet.cell(row=start + 5, column=4).value = 'группа'
    work_sheet.cell(row=start + 5, column=7).value = group.Name

    start_column = 1
    end_column = (len(courses.keys()) + 2) * 3
    work_sheet.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=end_column)
    work_sheet.merge_cells(start_row=2, start_column=start_column, end_row=2, end_column=end_column)
    work_sheet.merge_cells(start_row=3, start_column=start_column, end_row=3, end_column=end_column)

    work_sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)
    work_sheet.merge_cells(start_row=6, start_column=1, end_row=6, end_column=2)
    work_sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=2)

    work_sheet.cell(row=start, column=1).alignment = Alignment(horizontal='center', vertical='center')
    work_sheet.cell(row=start + 1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    work_sheet.cell(row=start + 2, column=1).alignment = Alignment(horizontal='center', vertical='center')
    work_sheet.cell(row=start + 4, column=3).alignment = Alignment(horizontal='center', vertical='center')
    work_sheet.cell(row=start + 5, column=3).alignment = Alignment(horizontal='center', vertical='center')
    return start + 7


def print_table_headers(work_sheet, courses, start):
    check_points = CheckPoint.objects.all()
    work_sheet.cell(row=start, column=1).value = '№'
    work_sheet.merge_cells(start_row=start, start_column=1, end_row=start+1, end_column=1)
    work_sheet.cell(row=start, column=2).value = 'Фамилия, имя, отчество'
    work_sheet.merge_cells(start_row=start, start_column=2, end_row=start + 1, end_column=2)
    work_sheet.cell(row=start, column=3).value = '№ зачетной книжки'
    work_sheet.merge_cells(start_row=start, start_column=3, end_row=start + 1, end_column=3)
    work_sheet.cell(row=start, column=1).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    work_sheet.cell(row=start, column=2).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    work_sheet.cell(row=start, column=3).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    col = 4
    for course_id in courses.keys():
        work_sheet.cell(row=start, column=col).value = courses[course_id]
        work_sheet.merge_cells(start_row=start, start_column=col, end_row=start, end_column=col+2)
        work_sheet.cell(row=start, column=col).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center', text_rotation=90)
        max_points = dict(CourseMaxPoints.objects.filter(course__id=course_id).values_list('checkpoint__id', 'max_point'))
        for check_point in check_points:
            max_point = str(max_points[check_point.id]) if check_point.id in max_points.keys() else ''
            work_sheet.cell(row=start + 1, column=col).value = check_point.name + ' max=' + max_point
            work_sheet.cell(row=start + 1, column=col).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center', text_rotation=90)
            col += 1
    work_sheet.cell(row=start, column=col).value = 'Пропущено всего/в т.ч. по уважительной причине'
    work_sheet.cell(row=start, column=col).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center',
                                                                 text_rotation=90)
    work_sheet.merge_cells(start_row=start, start_column=col, end_row=start, end_column=col + 2)
    work_sheet.column_dimensions['A'].width = 5
    work_sheet.column_dimensions['B'].width = 35
    work_sheet.column_dimensions['C'].width = 10
    max_column = (len(courses) + 2) * 3
    for column in get_column_interval(start=4, end=max_column):
        work_sheet.column_dimensions[column].width = 5
    work_sheet.row_dimensions[start].height = 80
    work_sheet.row_dimensions[start + 1].height = 80


def print_table_row(work_sheet, start, number, student, scores, courses, checkpoints):
    work_sheet.cell(row=start, column=1).value = number
    work_sheet.cell(row=start, column=2).value = student.FIO
    # work_sheet.cell(row=start, column=3).value = student.student_id
    col = 4
    courses_keys = list(courses.keys())
    for i in range(len(courses_keys)):
        for j in range(len(checkpoints)):
            work_sheet.cell(row=start, column=col + 3 * i + j).value = scores[(courses_keys[i], checkpoints[j])] if (courses_keys[i], checkpoints[j]) in scores.keys() else ''


def print_table_body(work_sheet, group, semester, courses, start):
    i = 1
    for sl in group.grouplist_set.filter(active=True).order_by('student__FIO'):
        scores = {(item[0],item[1]): item[2] for item in BRSpoints.objects.filter(student__id=sl.student.id, course__discipline_detail__semester__id=semester.id).values_list('course__id', 'checkpoint__id', 'points')}
        print_table_row(work_sheet, start, i, sl.student, scores, courses, CheckPoint.objects.values_list('id', flat=True))
        start += 1
        i += 1


def print_document_body(work_sheet, group, semester, courses, start):
    print_table_headers(work_sheet, courses, start)
    print_table_body(work_sheet, group, semester, courses, start + 2)


def export_group_points(group, semester):
    # объект
    wb = Workbook()

    # активный лист
    ws = wb.active

    # название страницы
    # ws = wb.create_sheet('первая страница', 0)
    ws.title = group.Name
    courses = dict(Course.objects.filter(group__id=group.id,
                                         discipline_detail__semester__id=semester.id).values_list('id', 'discipline_detail__discipline__Name'))
    row = 1
    row = print_headers(ws, group, semester, courses, row)
    row += 1
    print_document_body(ws, group, semester, courses, row)

    return wb


def print_exam_headers(work_sheet, data, start):
    period = EduPeriod.objects.get(active=True)
    work_sheet.cell(row=start, column=1).value = 'ФГАОУ ВО СЕВЕРО-ВОСТОЧНЫЙ ФЕДЕРАЛЬНЫЙ УНИВЕРСИТЕТ им. М.К. АММОСОВА'
    work_sheet.cell(row=start + 1, column=1).value = 'ИНСТИТУТ МАТЕМАТИКИ И ИНФОРМАТИКИ'
    work_sheet.cell(row=start + 2, column=1).value = 'ЖУРНАЛ ПРОМЕЖУТОЧНОЙ АТТЕСТАЦИИ за ' + str(period) + ' уч. год'
    work_sheet.cell(row=start + 4, column=1).value = 'Семестр'
    work_sheet.cell(row=start + 4, column=3).value = data['semester']
    work_sheet.cell(row=start + 5, column=1).value = 'Курс'
    work_sheet.cell(row=start + 6, column=1).value = 'Направление подготовки'
    work_sheet.cell(row=start + 5, column=3).value = data['group'].year
    work_sheet.cell(row=start + 6, column=3).value = data['group'].program.specialization.code + ' ' + data['group'].program.specialization.name
    work_sheet.cell(row=start + 5, column=4).value = 'группа'
    work_sheet.cell(row=start + 5, column=5).value = data['group'].Name

    start_column = 1
    end_column = len(data['courses'].keys()) + 3
    work_sheet.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=end_column)
    work_sheet.merge_cells(start_row=2, start_column=start_column, end_row=2, end_column=end_column)
    work_sheet.merge_cells(start_row=3, start_column=start_column, end_row=3, end_column=end_column)

    work_sheet.cell(row=start, column=1).alignment = Alignment(horizontal='center', vertical='center')
    work_sheet.cell(row=start + 1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    work_sheet.cell(row=start + 2, column=1).alignment = Alignment(horizontal='center', vertical='center')
    work_sheet.cell(row=start + 4, column=3).alignment = Alignment(horizontal='center', vertical='center')
    work_sheet.cell(row=start + 5, column=3).alignment = Alignment(horizontal='center', vertical='center')
    return start + 7


def print_exam_table_headers(work_sheet, courses, start):
    work_sheet.cell(row=start, column=1).value = '№'
    work_sheet.cell(row=start, column=2).value = 'Фамилия, имя, отчество'
    work_sheet.cell(row=start, column=3).value = '№ зачетной книжки'
    work_sheet.cell(row=start, column=1).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    work_sheet.cell(row=start, column=2).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    work_sheet.cell(row=start, column=3).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    col = 4
    for course_id in courses.keys():
        work_sheet.cell(row=start, column=col).value = courses[course_id]
        work_sheet.cell(row=start, column=col).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center', text_rotation=90)
        col += 1
    work_sheet.column_dimensions['A'].width = 5
    work_sheet.column_dimensions['B'].width = 45
    work_sheet.column_dimensions['C'].width = 10
    max_column = len(courses) + 3
    for column in get_column_interval(start=4, end=max_column):
        work_sheet.column_dimensions[column].width = 10
    work_sheet.row_dimensions[start].height = 90


def print_exam_table_row(work_sheet, start, number, row, courses):
    work_sheet.cell(row=start, column=1).value = number
    work_sheet.cell(row=start, column=2).value = row['fullname']
    col = 4
    courses_keys = list(courses.keys())
    scores = dict(row['scores'])
    for i in range(len(courses_keys)):
        if courses_keys[i] in scores:
            work_sheet.cell(row=start, column=col).value = scores[courses_keys[i]]
        col += 1


def print_exam_table_body(work_sheet, data, start):
    i = 1
    top = start
    for row in data['group_points']:
        print_exam_table_row(work_sheet, start, i, row, data['courses'])
        start += 1
        i += 1
    return start


def print_exam_document_body(work_sheet, data, start):
    top = start
    print_exam_table_headers(work_sheet, data['courses'], start)
    start = print_exam_table_body(work_sheet, data, start + 1)
    for row in work_sheet.iter_rows(min_row=top, min_col=1, max_row=start-1, max_col=len(data['courses']) + 3):
        for cell in row:
            work_sheet[cell.coordinate].border = Border(left=Side(border_style='thin',color='FF000000'),
                                                        right=Side(border_style='thin', color='FF000000'),
                                                        top=Side(border_style='thin', color='FF000000'),
                                                        bottom=Side(border_style='thin', color='FF000000'))
    return start


def export_exam_points(data):
    # объект
    wb = Workbook()
    ws = wb.active
    data['courses'] = dict(data['courses'])
    ws.title = data['group'].Name
    row = 1
    row = print_exam_headers(ws, data, row)
    row += 1
    row = print_exam_document_body(ws, data, row)

    ws.print_area = 'A1:' + ws.cell(row=row-1, column=len(data['courses'])+3).coordinate
    ws.page_setup = PrintPageSetup(worksheet=ws)
    ws.page_setup.paperSize = '9'
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToHeight = True
    ws.page_setup.fitToWidth = True
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return wb
